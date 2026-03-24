"""
App Rétro — Outil de supervision des déclarations DFX
Module 1 : Concaténation et rapport des déclarations DFX
"""

import streamlit as st
import openpyxl
import os
import re
import io
import json
import hashlib
import secrets
import pandas as pd
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ══════════════════════════════════════════════════════════════════════════════
#  CHARTE GRAPHIQUE BEAC
# ══════════════════════════════════════════════════════════════════════════════

def _inject_beac_css():
    """Injecte la feuille de style aux couleurs de la BEAC."""
    st.markdown("""
    <style>
    /* ── Palette BEAC ──────────────────────────────────────────────────────── */
    :root {
        --beac-bleu:    #003087;
        --beac-bleu2:   #00205B;
        --beac-or:      #C8A951;
        --beac-or-clair:#E8C96B;
        --beac-gris:    #F4F6FA;
        --beac-blanc:   #FFFFFF;
        --beac-texte:   #1A1A2E;
        --beac-bordure: #DDE3F0;
    }

    /* ── Fond général ──────────────────────────────────────────────────────── */
    .stApp {
        background-color: var(--beac-gris);
        font-family: 'Segoe UI', 'Helvetica Neue', Arial, sans-serif;
    }

    /* ── En-tête Streamlit (barre du haut) ─────────────────────────────────── */
    header[data-testid="stHeader"] {
        background-color: var(--beac-bleu2) !important;
        border-bottom: 3px solid var(--beac-or);
    }

    /* ── Sidebar ───────────────────────────────────────────────────────────── */
    section[data-testid="stSidebar"] {
        background: linear-gradient(180deg, var(--beac-bleu2) 0%, var(--beac-bleu) 100%) !important;
        border-right: 3px solid var(--beac-or);
    }
    section[data-testid="stSidebar"] * {
        color: var(--beac-blanc) !important;
    }
    section[data-testid="stSidebar"] .stRadio label {
        color: var(--beac-blanc) !important;
        font-weight: 500;
    }
    section[data-testid="stSidebar"] hr {
        border-color: var(--beac-or) !important;
        opacity: 0.5;
    }
    /* Radio button actif */
    section[data-testid="stSidebar"] .stRadio [data-testid="stMarkdownContainer"] p {
        color: var(--beac-blanc) !important;
    }
    section[data-testid="stSidebar"] input[type="radio"]:checked + div {
        border-color: var(--beac-or) !important;
    }

    /* ── Titres (h1, h2, h3) ────────────────────────────────────────────────── */
    h1, h2 {
        color: var(--beac-bleu) !important;
        border-bottom: 2px solid var(--beac-or);
        padding-bottom: 6px;
        margin-bottom: 16px;
    }
    h3 {
        color: var(--beac-bleu) !important;
    }

    /* ── Boutons primaires ─────────────────────────────────────────────────── */
    .stButton > button[kind="primary"],
    .stButton > button[data-testid="baseButton-primary"] {
        background: linear-gradient(135deg, var(--beac-bleu) 0%, var(--beac-bleu2) 100%) !important;
        color: var(--beac-blanc) !important;
        border: 2px solid var(--beac-or) !important;
        border-radius: 6px !important;
        font-weight: 600 !important;
        font-size: 15px !important;
        letter-spacing: 0.4px;
        transition: all 0.2s ease;
        box-shadow: 0 2px 8px rgba(0,48,135,0.25);
    }
    .stButton > button[kind="primary"]:hover {
        background: linear-gradient(135deg, var(--beac-or) 0%, var(--beac-or-clair) 100%) !important;
        color: var(--beac-bleu2) !important;
        border-color: var(--beac-bleu) !important;
        transform: translateY(-1px);
        box-shadow: 0 4px 14px rgba(200,169,81,0.4);
    }

    /* ── Boutons de téléchargement ─────────────────────────────────────────── */
    .stDownloadButton > button {
        background: linear-gradient(135deg, var(--beac-or) 0%, var(--beac-or-clair) 100%) !important;
        color: var(--beac-bleu2) !important;
        border: 2px solid var(--beac-bleu) !important;
        border-radius: 6px !important;
        font-weight: 700 !important;
        font-size: 14px !important;
        box-shadow: 0 2px 8px rgba(200,169,81,0.3);
        transition: all 0.2s ease;
    }
    .stDownloadButton > button:hover {
        background: linear-gradient(135deg, var(--beac-bleu) 0%, var(--beac-bleu2) 100%) !important;
        color: var(--beac-blanc) !important;
        border-color: var(--beac-or) !important;
        transform: translateY(-1px);
    }

    /* ── Métriques ─────────────────────────────────────────────────────────── */
    [data-testid="stMetric"] {
        background: var(--beac-blanc);
        border: 1px solid var(--beac-bordure);
        border-radius: 10px;
        padding: 14px 18px;
        border-left: 4px solid var(--beac-or) !important;
        box-shadow: 0 2px 8px rgba(0,48,135,0.08);
    }
    [data-testid="stMetricLabel"] {
        color: var(--beac-bleu) !important;
        font-weight: 600 !important;
        font-size: 13px !important;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }
    [data-testid="stMetricValue"] {
        color: var(--beac-bleu2) !important;
        font-size: 26px !important;
        font-weight: 700 !important;
    }

    /* ── Onglets (tabs) ────────────────────────────────────────────────────── */
    .stTabs [data-baseweb="tab-list"] {
        background-color: var(--beac-blanc);
        border-bottom: 2px solid var(--beac-bleu);
        border-radius: 6px 6px 0 0;
        gap: 4px;
        padding: 0 8px;
    }
    .stTabs [data-baseweb="tab"] {
        color: var(--beac-bleu) !important;
        font-weight: 600;
        padding: 8px 18px;
        border-radius: 6px 6px 0 0;
    }
    .stTabs [aria-selected="true"] {
        background-color: var(--beac-bleu) !important;
        color: var(--beac-blanc) !important;
        border-bottom: 3px solid var(--beac-or) !important;
    }

    /* ── Expanders ─────────────────────────────────────────────────────────── */
    .streamlit-expanderHeader {
        background-color: var(--beac-blanc) !important;
        border: 1px solid var(--beac-bordure) !important;
        border-left: 4px solid var(--beac-bleu) !important;
        border-radius: 6px !important;
        color: var(--beac-bleu) !important;
        font-weight: 600 !important;
        font-size: 14px !important;
    }
    .streamlit-expanderHeader:hover {
        border-left-color: var(--beac-or) !important;
        background-color: #EEF2FB !important;
    }
    .streamlit-expanderContent {
        border: 1px solid var(--beac-bordure);
        border-top: none;
        border-radius: 0 0 6px 6px;
        background-color: var(--beac-blanc);
        padding: 12px;
    }

    /* ── Containers ────────────────────────────────────────────────────────── */
    [data-testid="stVerticalBlock"] > [data-testid="element-container"] > div[style*="border"] {
        border-color: var(--beac-bordure) !important;
        border-radius: 10px !important;
        background: var(--beac-blanc);
    }

    /* ── Inputs / selectbox ────────────────────────────────────────────────── */
    .stTextInput > div > div > input,
    .stSelectbox > div > div {
        border-color: var(--beac-bleu) !important;
        border-radius: 6px !important;
    }
    .stTextInput > div > div > input:focus {
        border-color: var(--beac-or) !important;
        box-shadow: 0 0 0 2px rgba(200,169,81,0.2) !important;
    }

    /* ── DataFrames ────────────────────────────────────────────────────────── */
    [data-testid="stDataFrame"] {
        border-radius: 8px;
        overflow: hidden;
        border: 1px solid var(--beac-bordure);
        box-shadow: 0 2px 8px rgba(0,48,135,0.06);
    }

    /* ── Messages succès / warning / erreur ─────────────────────────────────── */
    [data-testid="stAlert"][data-type="success"] {
        border-left: 4px solid #28A745 !important;
        background-color: #F0FFF4 !important;
    }
    [data-testid="stAlert"][data-type="warning"] {
        border-left: 4px solid var(--beac-or) !important;
        background-color: #FFFBF0 !important;
    }
    [data-testid="stAlert"][data-type="error"] {
        border-left: 4px solid #DC3545 !important;
    }

    /* ── Dividers ──────────────────────────────────────────────────────────── */
    hr {
        border-color: var(--beac-or) !important;
        opacity: 0.4;
    }

    /* ── Spinner ───────────────────────────────────────────────────────────── */
    [data-testid="stSpinner"] {
        color: var(--beac-or) !important;
    }

    /* ── Caption ───────────────────────────────────────────────────────────── */
    [data-testid="stCaptionContainer"] p {
        color: #556080 !important;
        font-style: italic;
    }
    </style>
    """, unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
#  CONFIGURATION GLOBALE
# ══════════════════════════════════════════════════════════════════════════════

# Répertoire du script — fonctionne en local et sur Streamlit Cloud
WORKSPACE = os.path.dirname(os.path.abspath(__file__))
FICHIER_CODES_BANQUES = os.path.join(WORKSPACE, "Codes_banques.xlsx")
FICHIER_USERS        = os.path.join(WORKSPACE, "users.json")
# Code banque valide : 4-5 chiffres, code BIC (ex: UCMACMCX), ou tout identifiant ≥ 2 caractères
PATTERN_CODE_VALIDE = re.compile(r"^[\w\-']{2,}$", re.IGNORECASE)

# Paramètres spécifiques à chaque type de DFX
DFX_CONFIG = {
    "DFX_1200M": {
        "label"         : "DFX 1200M",
        "cellule_code"  : "D5",
        "ligne_entete"  : 7,   # ligne bleue en-tête
        "ligne_debut"   : 8,   # 1ère ligne de données
        "col_debut"     : 3,   # colonne C
        "feuille_ref"   : "Codes banques",
        "valider_code"  : True,
        "montant_cols"  : [10],
        "entetes"       : [
            "CODE BANQUE",
            "REFERENCES SWIFT",
            "MT",
            "DATE D'EMISSION DU MT",
            "DATE DE VALEUR DE MT",
            "CORRESPONDANT HORS CEMAC",
            "CODE BIC DU CORRESPONDANT HORS CEMAC",
            "CODE BIC DU CORRESPONDANT BEAC BENEFICIAIRE",
            "CENTRE - BEAC",
            "MONTANT EN DEVISE",
            "DEVISES",
            "DATE DU RAPATRIEMENT / REFERENCE DU MESSAGE SWIFT DU RAPATRIEMENT",
        ],
        "largeurs": {
            1: 15, 2: 22, 3: 10, 4: 20, 5: 20,
            6: 32, 7: 28, 8: 28, 9: 22, 10: 20,
            11: 12, 12: 38, 13: 28,
        },
    },
    "DFX_1401M": {
        "label"         : "DFX 1401M",
        "cellule_code"  : "D6",
        "ligne_entete"  : 8,
        "ligne_debut"   : 9,
        "col_debut"     : 3,
        "feuille_ref"   : "Listes banques 1401M",
        "valider_code"  : True,
        "montant_cols"  : [11],   # col 11 = MONTANT EN DEVISE (2 cols extra par rapport à l'ancien modèle)
        "entetes"       : [
            "CODE BANQUE",
            "REFERENCES SWIFT",
            "MT / MX",
            "DATE D'EMISSION DU MT / MX",
            "DATE DE VALEUR DU MT / MX",
            "CORRESPONDANT HORS CEMAC",
            "CODE BIC DU CORRESPONDANT HORS CEMAC",
            "CLIENT DE LA BANQUE / BENEFICIAIRE FINAL",       # col I — présente dans le modèle réel
            "CODE BIC DU CORRESPONDANT BEAC BENEFICIAIRE",
            "NUMERO DU COMPTE DETENU A LA BEAC",              # col K — présente dans le modèle réel
            "MONTANT EN DEVISE",
            "DEVISE",
            "DATE DU RAPATRIEMENT / REFERENCE DU MESSAGE SWIFT DU RAPATRIEMENT",
        ],
        "largeurs": {
            1: 15, 2: 22, 3: 10, 4: 20, 5: 20,
            6: 32, 7: 28, 8: 35, 9: 28, 10: 32,
            11: 20, 12: 12, 13: 38, 14: 28,
        },
    },
    "DFX_1500M": {
        "label"         : "DFX 1500M",
        "cellule_code"  : "D5",
        "ligne_entete"  : 7,
        "ligne_debut"   : 8,
        "col_debut"     : 3,
        "feuille_ref"   : "Codes banques",
        "valider_code"  : True,
        "montant_cols"  : [11, 14],
        "entetes"       : [
            "CODE BANQUE",
            "Date du message Swift reçu",
            "Type de message Swift reçu",
            "Référence du message Swift reçu",
            "Nom ou raison sociale du bénéficiaire",
            "Numéro de compte du bénéficiaire",
            "Devise du compte du bénéficiaire",
            "Personalité du bénéficiaire (physique / morale)",
            "Secteur d'activité du bénéficiaire",
            "Pays d'implantation du bénéficiaire",
            "Montant en devises de l'opération",
            "Devise de l'opération",
            "Cours de change appliqué",
            "Montant en FCFA de l'opération",
            "Correspondant hors CEMAC",
            "Code bic du correspondant hors CEMAC",
            "Nom ou raison sociale du donneur d'ordre",
            "Motif exact de l'opération",
        ],
        "largeurs": {
            1: 15,  2: 22,  3: 22,  4: 30,  5: 38,
            6: 30,  7: 22,  8: 32,  9: 32, 10: 32,
            11: 28, 12: 20, 13: 24, 14: 28, 15: 32,
            16: 30, 17: 38, 18: 38, 19: 30,
        },
    },
    # ── Variantes annuelles (Y) ─────────────────────────────────────────────
    "DFX_1200Y": {
        "label"         : "DFX 1200Y (Annuelle)",
        "cellule_code"  : "D5",
        "ligne_entete"  : 7,
        "ligne_debut"   : 8,
        "col_debut"     : 3,
        "feuille_ref"   : "Codes banques",
        "valider_code"  : True,
        "montant_cols"  : [10],
        "entetes"       : [
            "CODE BANQUE",
            "REFERENCES SWIFT",
            "MT / MX",
            "DATE D'EMISSION DU MT / MX",
            "DATE DE VALEUR DU MT / MX",
            "CORRESPONDANT HORS CEMAC",
            "CODE BIC DU CORRESPONDANT HORS CEMAC",
            "CODE BIC DU CORRESPONDANT BEAC BENEFICIAIRE",
            "CENTRE - BEAC",
            "MONTANT EN DEVISE",
            "DEVISE",
            "DATE DU RAPATRIEMENT / REFERENCE DU MESSAGE SWIFT DU RAPATRIEMENT",
        ],
        "largeurs": {
            1: 15, 2: 22, 3: 10, 4: 20, 5: 20,
            6: 32, 7: 28, 8: 28, 9: 22, 10: 20,
            11: 12, 12: 38, 13: 28,
        },
    },
    "DFX_1401Y": {
        "label"         : "DFX 1401Y (Annuelle)",
        "cellule_code"  : "D6",
        "ligne_entete"  : 8,
        "ligne_debut"   : 9,
        "col_debut"     : 3,
        "feuille_ref"   : "Listes banques 1401M",
        "valider_code"  : True,
        "montant_cols"  : [11],
        "entetes"       : [
            "CODE BANQUE",
            "REFERENCES SWIFT",
            "MT / MX",
            "DATE D'EMISSION DU MT / MX",
            "DATE DE VALEUR DU MT / MX",
            "CORRESPONDANT HORS CEMAC",
            "CODE BIC DU CORRESPONDANT HORS CEMAC",
            "CLIENT DE LA BANQUE / BENEFICIAIRE FINAL",
            "CODE BIC DU CORRESPONDANT BEAC BENEFICIAIRE",
            "NUMERO DU COMPTE DETENU A LA BEAC",
            "MONTANT EN DEVISE",
            "DEVISE",
            "DATE DU RAPATRIEMENT / REFERENCE DU MESSAGE SWIFT DU RAPATRIEMENT",
        ],
        "largeurs": {
            1: 15, 2: 22, 3: 10, 4: 20, 5: 20,
            6: 32, 7: 28, 8: 35, 9: 28, 10: 32,
            11: 20, 12: 12, 13: 38, 14: 28,
        },
    },
    "DFX_1500Y": {
        "label"         : "DFX 1500Y (Annuelle)",
        "cellule_code"  : "D5",
        "ligne_entete"  : 7,
        "ligne_debut"   : 8,
        "col_debut"     : 3,
        "feuille_ref"   : "Codes banques",
        "valider_code"  : True,
        "montant_cols"  : [11, 14],
        "entetes"       : [
            "CODE BANQUE",
            "Date du message Swift reçu",
            "Type de message Swift reçu",
            "Référence du message Swift reçu",
            "Nom ou raison sociale du bénéficiaire",
            "Numéro de compte du bénéficiaire",
            "Devise du compte du bénéficiaire",
            "Personalité du bénéficiaire (physique / morale)",
            "Secteur d'activité du bénéficiaire",
            "Pays d'implantation du bénéficiaire",
            "Montant en devises de l'opération",
            "Devise de l'opération",
            "Cours de change appliqué",
            "Montant en FCFA de l'opération",
            "Correspondant hors CEMAC",
            "Code bic du correspondant hors CEMAC",
            "Nom ou raison sociale du donneur d'ordre",
            "Motif exact de l'opération",
        ],
        "largeurs": {
            1: 15,  2: 22,  3: 22,  4: 30,  5: 38,
            6: 30,  7: 22,  8: 32,  9: 32, 10: 32,
            11: 28, 12: 20, 13: 24, 14: 28, 15: 32,
            16: 30, 17: 38, 18: 38, 19: 30,
        },
    },
}


# ══════════════════════════════════════════════════════════════════════════════
#  GESTION DES PROFILS UTILISATEURS
# ══════════════════════════════════════════════════════════════════════════════

# Modules accessibles par rôle
ROLES_MODULES = {
    "admin":           ["Concaténation DFX", "Taux de rétrocession", "Domiciliations Export"],
    "analyste_dfx":    ["Concaténation DFX", "Taux de rétrocession"],
    "superviseur_dom": ["Domiciliations Export"],
}

ROLES_LABELS = {
    "admin":           "Administrateur",
    "analyste_dfx":    "Analyste DFX",
    "superviseur_dom": "Superviseur DOM Export",
}

ROLES_BADGES = {
    "admin":           ("👑", "#C8A951", "#00205B"),
    "analyste_dfx":    ("🔬", "#27AE60", "#FFFFFF"),
    "superviseur_dom": ("🚢", "#2E86AB", "#FFFFFF"),
}


def _hash_password(password, salt=None):
    """Hache un mot de passe avec PBKDF2-SHA256 + sel aléatoire."""
    if salt is None:
        salt = secrets.token_hex(16)
    hashed = hashlib.pbkdf2_hmac(
        "sha256", password.encode("utf-8"), salt.encode("utf-8"), 200_000
    ).hex()
    return hashed, salt


def _verifier_password(password, stored_hash, salt):
    """Vérifie un mot de passe contre le hash stocké."""
    computed, _ = _hash_password(password, salt)
    return computed == stored_hash


def _charger_users():
    """Charge le fichier users.json. Le crée avec les comptes par défaut si absent."""
    if not os.path.exists(FICHIER_USERS):
        return _creer_users_defaut()
    with open(FICHIER_USERS, "r", encoding="utf-8") as f:
        return json.load(f)


def _sauvegarder_users(data):
    """Sauvegarde le dict utilisateurs dans users.json."""
    with open(FICHIER_USERS, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def _creer_users_defaut():
    """Crée le fichier users.json avec 3 comptes par défaut."""
    comptes_defaut = [
        ("admin",      "Admin@BEAC2026",  "admin",           "Administrateur BEAC"),
        ("analyste",   "DFX@2026",        "analyste_dfx",    "Analyste DFX"),
        ("dom_export", "Export@2026",     "superviseur_dom", "Superviseur DOM Export"),
    ]
    data = {"users": []}
    for username, password, role, display_name in comptes_defaut:
        hashed, salt = _hash_password(password)
        data["users"].append({
            "username":             username,
            "password_hash":        hashed,
            "salt":                 salt,
            "display_name":         display_name,
            "role":                 role,
            "must_change_password": True,
        })
    _sauvegarder_users(data)
    return data


def _trouver_utilisateur(data, username):
    """Retourne le dict utilisateur ou None."""
    for user in data["users"]:
        if user["username"] == username:
            return user
    return None


def _page_login():
    """Affiche la page de connexion centrée."""
    # Masque la sidebar sur la page de login
    st.markdown("""
    <style>
    [data-testid="stSidebar"] { display: none !important; }
    </style>
    """, unsafe_allow_html=True)

    _do_rerun = False
    col_l, col_c, col_r = st.columns([1, 1.2, 1])
    with col_c:
        # Carte login
        st.markdown("""
        <div style="
            background: linear-gradient(145deg, #002060 0%, #003087 60%, #004CB3 100%);
            border-radius: 18px;
            padding: 40px 36px 32px;
            border: 1px solid rgba(200,169,81,0.4);
            box-shadow: 0 8px 40px rgba(0,32,96,0.35);
            text-align: center;
            margin-top: 40px;
        ">
            <div style="font-size: 58px; margin-bottom: 8px;">🏦</div>
            <div style="font-size: 22px; font-weight: 900; color: #C8A951;
                        letter-spacing: 3px; text-transform: uppercase;">BEAC</div>
            <div style="font-size: 13px; color: #AAC0E0; margin: 4px 0 20px;
                        letter-spacing: 1px; text-transform: uppercase;">
                Application de Supervision DFX
            </div>
            <div style="border-top: 1px solid rgba(200,169,81,0.3);
                        padding-top: 16px; color: #C8D8F0; font-size: 13px;">
                Veuillez vous identifier pour accéder aux modules
            </div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("<div style='height:20px'></div>", unsafe_allow_html=True)

        username = st.text_input("👤 Identifiant", placeholder="Votre identifiant")
        password = st.text_input("🔒 Mot de passe", type="password", placeholder="Votre mot de passe")

        if st.button("🔓 Se connecter", use_container_width=True, type="primary"):
            if not username or not password:
                st.error("Veuillez renseigner l'identifiant et le mot de passe.")
                return

            data = _charger_users()
            user = _trouver_utilisateur(data, username.strip())

            if user is None or not _verifier_password(password, user["password_hash"], user["salt"]):
                st.error("❌ Identifiant ou mot de passe incorrect.")
                return

            # Connexion réussie
            st.session_state["authenticated"]  = True
            st.session_state["username"]        = user["username"]
            st.session_state["display_name"]    = user["display_name"]
            st.session_state["role"]            = user["role"]
            st.session_state["must_change_pwd"] = user.get("must_change_password", False)
            _do_rerun = True

        # Première connexion : afficher les identifiants par défaut une seule fois
        if not os.path.exists(FICHIER_USERS):
            with st.expander("ℹ️ Identifiants par défaut (première connexion)"):
                st.markdown("""
                | Identifiant | Mot de passe | Rôle |
                |---|---|---|
                | `admin` | `Admin@BEAC2026` | Administrateur |
                | `analyste` | `DFX@2026` | Analyste DFX |
                | `dom_export` | `Export@2026` | Superviseur DOM |
                """)
    # Déclenchement du rerun HORS contexte column (évite l'erreur removeChild)
    if _do_rerun:
        st.rerun()


def _page_changer_password():
    """Formulaire de changement de mot de passe (obligatoire à la 1ère connexion)."""
    st.warning("⚠️ Changez votre mot de passe avant de continuer.")
    st.markdown("### 🔑 Nouveau mot de passe")

    new_pwd   = st.text_input("Nouveau mot de passe", type="password")
    confirm   = st.text_input("Confirmer le mot de passe", type="password")

    if st.button("✅ Valider le changement", type="primary"):
        if len(new_pwd) < 8:
            st.error("Le mot de passe doit comporter au moins 8 caractères.")
            return
        if new_pwd != confirm:
            st.error("Les deux mots de passe ne correspondent pas.")
            return

        data = _charger_users()
        user = _trouver_utilisateur(data, st.session_state["username"])
        if user:
            new_hash, new_salt = _hash_password(new_pwd)
            user["password_hash"]       = new_hash
            user["salt"]                = new_salt
            user["must_change_password"] = False
            _sauvegarder_users(data)
            st.session_state["must_change_pwd"] = False
            st.success("✅ Mot de passe changé avec succès !")
            st.rerun()


def _sidebar_gestion_users():
    """Interface admin de gestion des utilisateurs (dans un expander)."""
    with st.expander("⚙️ Gestion des utilisateurs", expanded=False):
        data = _charger_users()

        st.markdown("#### Utilisateurs enregistrés")
        for u in data["users"]:
            role_lbl = ROLES_LABELS.get(u["role"], u["role"])
            ico, bg, fg = ROLES_BADGES.get(u["role"], ("👤", "#888", "#FFF"))
            st.markdown(f"""
            <div style="background:rgba(0,48,135,0.08);border-radius:8px;
                        padding:8px 12px;margin:4px 0;font-size:13px;
                        border-left:3px solid {bg};">
                <strong>{u['display_name']}</strong>
                &nbsp;<code>{u['username']}</code>
                &nbsp;<span style="background:{bg};color:{fg};
                                   padding:2px 8px;border-radius:10px;
                                   font-size:11px;">{ico} {role_lbl}</span>
            </div>
            """, unsafe_allow_html=True)

        st.markdown("---")
        st.markdown("#### Ajouter un utilisateur")
        new_user    = st.text_input("Identifiant", key="admin_new_user")
        new_name    = st.text_input("Nom affiché", key="admin_new_name")
        new_role    = st.selectbox("Rôle", options=list(ROLES_LABELS.keys()),
                                   format_func=lambda r: ROLES_LABELS[r],
                                   key="admin_new_role")
        new_pwd_adm = st.text_input("Mot de passe initial", type="password",
                                    key="admin_new_pwd")
        if st.button("➕ Créer le compte", key="admin_create_btn"):
            if not new_user or not new_pwd_adm:
                st.error("Identifiant et mot de passe requis.")
            elif _trouver_utilisateur(data, new_user.strip()):
                st.error("Cet identifiant existe déjà.")
            else:
                hashed, salt = _hash_password(new_pwd_adm)
                data["users"].append({
                    "username":             new_user.strip(),
                    "password_hash":        hashed,
                    "salt":                 salt,
                    "display_name":         new_name.strip() or new_user.strip(),
                    "role":                 new_role,
                    "must_change_password": True,
                })
                _sauvegarder_users(data)
                st.success(f"✅ Compte « {new_user} » créé avec succès.")
                st.session_state["_gestion_rerun"] = True

        st.markdown("---")
        st.markdown("#### Réinitialiser un mot de passe")
        usernames   = [u["username"] for u in data["users"]]
        sel_user    = st.selectbox("Utilisateur", options=usernames,
                                   key="admin_reset_user")
        reset_pwd   = st.text_input("Nouveau mot de passe", type="password",
                                    key="admin_reset_pwd")
        if st.button("🔄 Réinitialiser", key="admin_reset_btn"):
            if not reset_pwd:
                st.error("Entrez un nouveau mot de passe.")
            else:
                target = _trouver_utilisateur(data, sel_user)
                if target:
                    hashed, salt = _hash_password(reset_pwd)
                    target["password_hash"]       = hashed
                    target["salt"]                = salt
                    target["must_change_password"] = True
                    _sauvegarder_users(data)
                    st.success(f"✅ Mot de passe de « {sel_user} » réinitialisé.")


# ══════════════════════════════════════════════════════════════════════════════
#  FONCTIONS UTILITAIRES
# ══════════════════════════════════════════════════════════════════════════════

def lire_codes_banques_reference(type_dfx: str, fichier_ref_src=None) -> pd.DataFrame:
    """
    Lit la liste exhaustive des banques depuis le fichier de référence.
    Accepte un chemin str OU un UploadedFile/BytesIO.
    Retourne un DataFrame normalisé avec colonnes : CODE, NOM
    """
    src = fichier_ref_src or FICHIER_CODES_BANQUES
    if isinstance(src, str):
        src = src.replace("\\", "/").strip()
        if not os.path.exists(src):
            st.warning(f"Fichier de référence introuvable : `{src}`")
            return pd.DataFrame(columns=["CODE", "NOM"])
        data = src
    else:
        data = io.BytesIO(src.getvalue() if hasattr(src, "getvalue") else src.read())

    feuille = DFX_CONFIG[type_dfx]["feuille_ref"]
    try:
        xf = pd.ExcelFile(data)
        # Essaye la feuille configée, sinon prend la première disponible
        sheet = feuille if feuille in xf.sheet_names else xf.sheet_names[0]
        # Lecture avec pandas — dtype str pour éviter les entiers sur les codes
        df_raw = pd.read_excel(xf, sheet_name=sheet, header=0, dtype=str)
        df_raw = df_raw.dropna(how="all").reset_index(drop=True)

        # Nettoyage des valeurs
        df_raw = df_raw.apply(lambda col: col.str.strip() if col.dtype == object else col)

        # ── Détection robuste des colonnes CODE et NOM ──────────────────────
        # Stratégie : chercher d'abord CODE dans le NOM de colonne,
        # puis dans les VALEURS si non trouvé (cas Unnamed: x)
        col_code, col_nom = None, None

        # 1er passage : colonnes nommées explicitement
        for c in df_raw.columns:
            cl = str(c).lower()
            if col_code is None and "code" in cl:
                col_code = c
            if col_nom is None and any(k in cl for k in ("nom", "libel", "raison", "denom")) \
                    and "code" not in cl:
                col_nom = c

        # 2e passage : chercher dans les valeurs si colonnes non identifiées
        # (cas "Unnamed: x" où les vraies étiquettes sont dans la 1ère ligne de données)
        if col_code is None or col_nom is None:
            for c in df_raw.columns:
                if str(c).lower().startswith("unnamed"):
                    series_cl = df_raw[c].dropna().astype(str)
                    # Colonne CODE = contient des valeurs numériques à 4-5 chiffres ou BIC
                    if col_code is None:
                        num_valide = series_cl.str.match(r'^\d{4,6}(\.0)?$|^[A-Z]{4}\d{2}[A-Z]{2}').sum()
                        if num_valide >= len(series_cl) * 0.5 and len(series_cl) > 0:
                            col_code = c
                    # Colonne NOM = contient strings longs (noms de banques)
                    if col_nom is None:
                        long_str = series_cl.str.len().median()
                        if long_str and long_str > 8 and c != col_code:
                            col_nom = c

        # Repli final sur les 2 premières colonnes non-vides
        cols_non_vides = [c for c in df_raw.columns
                          if df_raw[c].dropna().astype(str).str.strip().ne("").any()]
        if col_code is None and cols_non_vides:
            col_code = cols_non_vides[0]
        if col_nom is None:
            for c in cols_non_vides:
                if c != col_code:
                    col_nom = c
                    break

        def _norm_code_ref(s: str) -> str:
            """Normalise un code banque : '10001.0' → '10001', 'CITIGB2L' intact."""
            s = s.strip()
            try:
                return str(int(float(s)))
            except (ValueError, OverflowError):
                return s.upper()

        df_result = pd.DataFrame()
        df_result["CODE"] = (df_raw[col_code].fillna("").astype(str)
                              .str.strip().apply(_norm_code_ref))
        df_result["NOM"]  = (df_raw[col_nom].fillna("").astype(str).str.strip()
                              if col_nom else "")

        # Supprimer les lignes sans code ou avec code parasite
        df_result = df_result[
            df_result["CODE"].ne("") &
            ~df_result["CODE"].isin(["NAN", "NONE", "CODE", "CODES"])
        ].reset_index(drop=True)
        return df_result

    except Exception as e:
        st.warning(f"Impossible de lire la feuille '{feuille}' : {e}")
        return pd.DataFrame(columns=["CODE", "NOM"])


def _lire_global_noms(fichier_ref_src=None) -> dict:
    """
    Lit TOUTES les feuilles du fichier référentiel et retourne un dict {CODE: NOM} unifié.
    Utilisé uniquement pour l'affichage des noms (ne remplace pas la liste de référence
    DFX-spécifique utilisée pour la logique déclarants / non-déclarants).
    """
    src = fichier_ref_src or FICHIER_CODES_BANQUES
    if isinstance(src, str):
        src = src.replace("\\", "/").strip()
        if not os.path.exists(src):
            return {}
        data = src
    else:
        data = io.BytesIO(src.getvalue() if hasattr(src, "getvalue") else src.read())

    map_noms: dict = {}
    try:
        xf = pd.ExcelFile(data)
        for sheet in xf.sheet_names:
            df = pd.read_excel(xf, sheet_name=sheet, header=0, dtype=str).dropna(how="all")
            df = df.apply(lambda c: c.str.strip() if c.dtype == object else c)
            cols = list(df.columns)
            col_code, col_nom = None, None
            for c in cols:
                cl = str(c).lower()
                if col_code is None and "code" in cl:
                    col_code = c
                if col_nom is None and any(k in cl for k in ("nom", "libel", "raison", "denom", "banque")) \
                        and "code" not in cl:
                    col_nom = c
            # Fallback : colonnes non nommées explicitement
            cols_ok = [c for c in cols
                       if df[c].dropna().astype(str).str.strip().ne("").any()]
            if col_code is None and cols_ok:
                col_code = cols_ok[0]
            if col_nom is None:
                for c in cols_ok:
                    if c != col_code:
                        col_nom = c
                        break
            if col_code is None or col_nom is None:
                continue
            for _, row in df[[col_code, col_nom]].iterrows():
                code_raw = str(row[col_code]).strip()
                nom_raw  = str(row[col_nom]).strip()
                if not code_raw or code_raw.upper() in ("NAN", "NONE", "CODE", "CODES", ""):
                    continue
                try:
                    code_norm = str(int(float(code_raw)))
                except (ValueError, OverflowError):
                    code_norm = code_raw.upper()
                if nom_raw and nom_raw.upper() not in ("NAN", "NONE", ""):
                    map_noms.setdefault(code_norm, nom_raw)  # premier trouvé conservé
    except Exception:
        pass
    return map_noms


def extraire_code_banque(ws, cellule: str, valider_code: bool):
    """
    Extrait le code banque depuis la cellule indiquée.
    Accepte : codes à 4-5 chiffres, codes BIC (ex: UCMACMCX), noms de banque.
    Retourne (code, erreur_msg). Si invalide, code=None.
    """
    valeur = ws[cellule].value
    if valeur is None or str(valeur).strip() == "":
        return None, f"cellule {cellule} vide"
    # Normaliser : openpyxl retourne parfois 10001.0 (float) → '10001'
    code = str(valeur).strip()
    try:
        code = str(int(float(code)))
    except (ValueError, OverflowError):
        code = code.strip()
    if valider_code and not PATTERN_CODE_VALIDE.match(code):
        return None, f"valeur '{code}' en {cellule} ne ressemble pas à un code banque"
    return code, None


def concatener_dfx(src, type_dfx: str) -> dict:
    """
    Extrait les données de chaque fichier Excel selon la configuration du type DFX.
    `src` peut être :
      - un str (chemin de dossier) → parcourt le dossier
      - une list d'UploadedFile    → itère sur les fichiers uploadés

    Retourne un dictionnaire :
      {
        "donnees"          : list[dict],   # lignes extraites
        "fichiers_ok"      : list[str],    # fichiers traités avec succès
        "codes_ok"         : set[str],     # codes banques valides trouvés
        "fichiers_mal_struct": list[dict], # {fichier, raison}
        "total_lignes"     : int,
      }
    """
    cfg = DFX_CONFIG[type_dfx]

    # Construction de la liste itérable : (nom_fichier, source_wb)
    if isinstance(src, str):
        fichiers_iter = [
            (f, os.path.join(src, f))
            for f in os.listdir(src) if f.endswith((".xlsx", ".xlsm"))
        ]
    else:
        fichiers_iter = [(f.name, io.BytesIO(f.getvalue())) for f in src]

    donnees            = []
    fichiers_ok        = []
    codes_ok           = set()
    fichiers_mal_struct = []
    total_lignes       = 0

    for nom_fichier, wb_src in fichiers_iter:
        try:
            wb = openpyxl.load_workbook(wb_src, data_only=True)
            ws = wb.active

            code_banque, err = extraire_code_banque(
                ws, cfg["cellule_code"], cfg["valider_code"]
            )
            if code_banque is None:
                fichiers_mal_struct.append({"Fichier": nom_fichier, "Raison": err})
                wb.close()
                continue

            max_row = ws.max_row
            max_col = ws.max_column

            if max_row < cfg["ligne_debut"]:
                fichiers_mal_struct.append({
                    "Fichier": nom_fichier,
                    "Raison" : f"aucune donnée (max_row={max_row}, début attendu={cfg['ligne_debut']})"
                })
                wb.close()
                continue

            lignes_fichier = 0
            for row in range(cfg["ligne_debut"], max_row + 1):
                ligne_data = []
                nb_vals    = 0
                for col in range(cfg["col_debut"], max_col + 1):
                    v = ws.cell(row=row, column=col).value
                    ligne_data.append(v)
                    if v is not None and str(v).strip() != "":
                        nb_vals += 1

                if nb_vals == 0:
                    continue  # ligne entièrement vide

                donnees.append({
                    "code_banque"   : code_banque,
                    "data"          : ligne_data,
                    "fichier_source": nom_fichier,
                })
                lignes_fichier += 1

            total_lignes += lignes_fichier
            fichiers_ok.append(nom_fichier)
            codes_ok.add(code_banque)
            wb.close()

        except Exception as e:
            fichiers_mal_struct.append({"Fichier": nom_fichier, "Raison": str(e)})

    return {
        "donnees"            : donnees,
        "fichiers_ok"        : fichiers_ok,
        "codes_ok"           : codes_ok,
        "fichiers_mal_struct": fichiers_mal_struct,
        "total_lignes"       : total_lignes,
    }


def construire_fichier_excel(donnees: list, entetes: list, largeurs: dict,
                              montant_cols: list, type_dfx: str,
                              sheets_rapport: list = None) -> bytes:
    """
    Construit le fichier Excel consolidé en mémoire et retourne les bytes.
    sheets_rapport : liste optionnelle de (nom_feuille, DataFrame, couleur_hex_bg)
                     Chaque feuille est ajoutée après la feuille de données principale.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Données {type_dfx}"

    entetes_complets = entetes + ["Fichier Source"]

    hdr_font  = Font(bold=True, size=11, color="FFFFFF")
    hdr_fill  = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border    = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    for col_idx, titre in enumerate(entetes_complets, start=1):
        cell           = ws.cell(row=1, column=col_idx)
        cell.value     = titre
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        cell.border    = border

    ligne = 2
    for item in donnees:
        # Col 1 : code banque
        c        = ws.cell(row=ligne, column=1, value=item["code_banque"])
        c.border = border

        for col_idx, valeur in enumerate(item["data"], start=2):
            c        = ws.cell(row=ligne, column=col_idx, value=valeur)
            c.border = border
            if isinstance(valeur, datetime):
                c.number_format = "DD/MM/YYYY"
            elif isinstance(valeur, (int, float)) and col_idx in montant_cols:
                c.number_format = "#,##0.00"

        # Dernière col : fichier source
        c        = ws.cell(row=ligne, column=len(entetes_complets), value=item["fichier_source"])
        c.border = border
        ligne   += 1

    for col_idx, larg in largeurs.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = larg

    ws.freeze_panes    = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── Feuilles rapport récapitulatif (déclarants / non-déclarants / rejets) ────
    if sheets_rapport:
        for sheet_name, df_sheet, bg_color in sheets_rapport:
            if df_sheet is None or (hasattr(df_sheet, "empty") and df_sheet.empty):
                continue
            ws_r      = wb.create_sheet(sheet_name[:31])
            rpt_fill  = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            rpt_font  = Font(bold=True, size=10, color="FFFFFF")
            rpt_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for ci, col_name in enumerate(df_sheet.columns, 1):
                cell           = ws_r.cell(row=1, column=ci, value=str(col_name))
                cell.font      = rpt_font
                cell.fill      = rpt_fill
                cell.alignment = rpt_align
                cell.border    = border
            for ri, row_data in enumerate(df_sheet.itertuples(index=False), 2):
                for ci, val in enumerate(row_data, 1):
                    cell        = ws_r.cell(row=ri, column=ci, value=val)
                    cell.border = border
                    if isinstance(val, float):
                        cell.number_format = "#,##0.00"
                        cell.alignment     = Alignment(horizontal="right", vertical="center")
                    elif isinstance(val, int):
                        cell.number_format = "#,##0"
                        cell.alignment     = Alignment(horizontal="right", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    if isinstance(val, str):
                        if "Déclarée" in val and "Non" not in val:
                            cell.font = Font(color="1F6B3E", bold=True, size=10)
                        elif "Non déclarée" in val:
                            cell.font = Font(color="8B0000", bold=True, size=10)
            ws_r.freeze_panes = "A2"
            for ci in range(1, len(df_sheet.columns) + 1):
                col_vals = df_sheet.iloc[:, ci - 1].fillna("").astype(str).tolist()
                max_len  = max([len(str(df_sheet.columns[ci - 1]))] + [len(v) for v in col_vals])
                ws_r.column_dimensions[
                    openpyxl.utils.get_column_letter(ci)
                ].width = min(max(max_len + 2, 12), 52)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
#  MODULE 1 : CONCATÉNATION DFX
# ══════════════════════════════════════════════════════════════════════════════

def module_concatenation():
    # ── 1. Paramètres ─────────────────────────────────────────────────────────
    with st.container(border=True):
        st.subheader("⚙️ Paramètres")

        col1, col2 = st.columns(2)
        with col1:
            type_dfx = st.selectbox(
                "Type de déclaration",
                options=list(DFX_CONFIG.keys()),
                format_func=lambda k: DFX_CONFIG[k]["label"],
            )
        with col2:
            nom_fichier_sortie = st.text_input(
                "Nom du fichier de sortie",
                value=f"{type_dfx}_Consolide.xlsx",
                help="Le fichier sera téléchargeable après la concaténation.",
            )

        # ── ZONE D'IMPORT PRINCIPALE ─────────────────────────────────────────
        st.markdown("""
        <div style="background:linear-gradient(135deg,#003087 0%,#00205B 100%);
                    border-radius:10px;border:2px solid #C8A951;
                    padding:16px 20px;margin:14px 0 8px 0;">
            <div style="display:flex;align-items:center;gap:14px;">
                <div style="background:#C8A951;border-radius:50%;width:40px;height:40px;
                            display:flex;align-items:center;justify-content:center;
                            font-size:20px;font-weight:900;color:#00205B;flex-shrink:0;">1</div>
                <div>
                    <div style="font-size:16px;font-weight:800;color:#FFFFFF;letter-spacing:0.3px;">
                        Fichiers DFX individuels des banques
                        <span style="font-size:12px;font-weight:600;color:#FF8C8C;
                                     background:rgba(255,100,100,0.2);border-radius:4px;
                                     padding:1px 7px;margin-left:8px;">OBLIGATOIRE</span>
                    </div>
                    <div style="font-size:12.5px;color:#9FBBDE;margin-top:4px;line-height:1.5;">
                        Importez <b style="color:#FFD97E;">un ou plusieurs</b> fichiers Excel DFX
                        — chaque fichier = <b style="color:#FFD97E;">une banque</b>.
                        Formats acceptés&nbsp;: <code style="color:#C8D8F0;">.xlsx</code>
                        &nbsp;/&nbsp;<code style="color:#C8D8F0;">.xlsm</code>
                    </div>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)
        fichiers_concat = st.file_uploader(
            "📂  Glissez-déposez vos fichiers DFX ici, ou cliquez pour parcourir",
            type=["xlsx", "xlsm"],
            accept_multiple_files=True,
            key="up_fichiers_concat",
            help="Exemples : CITIGALX_1770366278060.xlsm, BGFIGALX_1_DFX_NEW_MODELE_300125_V0.xlsx",
        )
        if fichiers_concat:
            noms = ", ".join(f.name for f in fichiers_concat[:4])
            extra = f" (+{len(fichiers_concat) - 4} autres)" if len(fichiers_concat) > 4 else ""
            st.success(f"✅  **{len(fichiers_concat)} fichier(s) chargé(s)** : {noms}{extra}")
        else:
            st.warning("⚠️  Aucun fichier sélectionné — importez vos fichiers DFX pour continuer.")

        # ── FICHIER DE RÉFÉRENCE ──────────────────────────────────────────────
        st.markdown("""
        <div style="background:linear-gradient(135deg,#3D2B00 0%,#2A1E00 100%);
                    border-radius:10px;border:2px solid #C8A951;
                    padding:14px 20px;margin:14px 0 8px 0;">
            <div style="display:flex;align-items:center;gap:14px;">
                <div style="background:rgba(200,169,81,0.25);border:2px solid #C8A951;border-radius:50%;
                            width:40px;height:40px;display:flex;align-items:center;justify-content:center;
                            font-size:20px;font-weight:900;color:#FFD97E;flex-shrink:0;">2</div>
                <div>
                    <div style="font-size:15px;font-weight:700;color:#FFD97E;">
                        Fichier de référence des codes banques
                        <span style="font-size:11px;font-weight:500;color:#C8A951;
                                     background:rgba(200,169,81,0.15);border-radius:4px;
                                     padding:1px 7px;margin-left:6px;">optionnel</span>
                    </div>
                    <div style="font-size:12px;color:#D4B870;margin-top:3px;">
                        Fichier Excel listant tous les <b>codes et noms des banques agréées</b>
                        — permet d'identifier les non-déclarants.
                        Ex.&nbsp;: <code style="color:#FFD97E;">Codes_banques_régul_Jan26.xlsx</code>
                    </div>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)
        up_ref_concat = st.file_uploader(
            "📋  Référentiel bancaire (Codes_banques_régul_Jan26.xlsx ou équivalent)",
            type=["xlsx", "xlsm"],
            key="up_ref_concat",
            help="Colonnes attendues : CODE BANQUE | NOM (ou libellé/dénomination). Permet d'identifier les non-déclarants.",
        )
        if up_ref_concat:
            st.success(f"✅  Référentiel chargé : **{up_ref_concat.name}**")

    # ── 2. Bouton de lancement ─────────────────────────────────────────────────
    if st.button("🚀 Lancer la concaténation", type="primary", use_container_width=True):

        # Résolution du fichier de référence
        src_ref = up_ref_concat if up_ref_concat is not None else None

        # Vérifications préalables
        if not fichiers_concat:
            st.error("❌ Veuillez importer au moins un fichier Excel à concaténer.")
            return

        # Lancement
        with st.spinner(f"Concaténation des {len(fichiers_concat)} fichiers en cours…"):
            resultat = concatener_dfx(fichiers_concat, type_dfx)

        cfg = DFX_CONFIG[type_dfx]

        # ── 3. Métriques rapides ───────────────────────────────────────────────
        st.divider()
        st.subheader("📊 Résultats")

        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Fichiers importés",   len(fichiers_concat))
        m2.metric("Fichiers traités ✅",  len(resultat["fichiers_ok"]))
        m3.metric("Lignes extraites",     resultat["total_lignes"])
        m4.metric("Fichiers rejetés ⚠️",  len(resultat["fichiers_mal_struct"]))

        # ── Chargement du référentiel anticipé (inclus dans le fichier consolidé) ──
        df_ref = lire_codes_banques_reference(type_dfx, src_ref)
        if not df_ref.empty:
            df_ref["CODE"] = df_ref["CODE"].str.upper().str.strip()
        codes_ref      = set(df_ref["CODE"].tolist()) if not df_ref.empty else set()
        codes_declares = {c.upper().strip() for c in resultat["codes_ok"]}

        # Construction des DataFrames pour les feuilles rapport dans le consolidé
        if codes_declares:
            if not df_ref.empty:
                _df_dec = df_ref[df_ref["CODE"].isin(codes_declares)][["CODE", "NOM"]].copy()
                _hors   = codes_declares - codes_ref
                if _hors:
                    _gnoms_sheets = _lire_global_noms(src_ref)
                    _df_dec = pd.concat([
                        _df_dec,
                        pd.DataFrame([
                            {"CODE": c, "NOM": _gnoms_sheets.get(c, "⚠️ Non répertoriée")}
                            for c in sorted(_hors)
                        ])
                    ], ignore_index=True)
            else:
                _df_dec = pd.DataFrame([{"CODE": c, "NOM": "—"} for c in sorted(codes_declares)])
            _df_dec = _df_dec.sort_values("CODE").reset_index(drop=True)
            _df_dec.columns = ["Code banque", "Nom banque"]
            _df_dec.insert(2, "Statut", "✅ Déclarée")
        else:
            _df_dec = pd.DataFrame(columns=["Code banque", "Nom banque", "Statut"])

        _codes_non_dec = codes_ref - codes_declares
        if not df_ref.empty and _codes_non_dec:
            _df_non_dec = df_ref[df_ref["CODE"].isin(_codes_non_dec)][["CODE", "NOM"]].copy()
            _df_non_dec.columns = ["Code banque", "Nom banque"]
            _df_non_dec.insert(2, "Statut", "❌ Non déclarée")
            _df_non_dec = _df_non_dec.sort_values("Code banque").reset_index(drop=True)
        else:
            _df_non_dec = pd.DataFrame(columns=["Code banque", "Nom banque", "Statut"])

        _df_rej = (pd.DataFrame(resultat["fichiers_mal_struct"])
                   if resultat["fichiers_mal_struct"]
                   else pd.DataFrame(columns=["Fichier", "Raison"]))

        _sheets_rapport = [
            ("Déclarants ✅",     _df_dec,     "1F6B3E"),
            ("Non-déclarants ❌", _df_non_dec, "8B0000"),
            ("Fichiers rejetés",  _df_rej,     "C8A951"),
        ]

        # ── 4. Téléchargement du fichier consolidé ─────────────────────────────
        if resultat["donnees"]:
            excel_bytes = construire_fichier_excel(
                donnees        = resultat["donnees"],
                entetes        = cfg["entetes"],
                largeurs       = cfg["largeurs"],
                montant_cols   = cfg["montant_cols"],
                type_dfx       = type_dfx,
                sheets_rapport = _sheets_rapport,
            )
            st.download_button(
                label     = f"⬇️ Télécharger {nom_fichier_sortie}",
                data      = excel_bytes,
                file_name = nom_fichier_sortie,
                mime      = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        else:
            st.warning("⚠️ Aucune donnée extraite — le fichier consolidé n'a pas pu être généré.")

        # ── 5. RAPPORT ────────────────────────────────────────────────────────
        st.divider()
        st.subheader(f"📋 Rapport de déclaration — {cfg['label']}")
        st.caption(f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}")

        # (df_ref, codes_ref, codes_declares déjà calculés ci-dessus)

        # ── 5a. Banques ayant déclaré ──────────────────────────────────────────
        with st.expander(f"✅ Banques ayant déclaré ({len(codes_declares)})", expanded=True):
            if codes_declares:
                if not df_ref.empty:
                    df_declares = df_ref[df_ref["CODE"].isin(codes_declares)][["CODE", "NOM"]].copy()
                    # Banques hors référentiel mais avec code valide
                    hors_ref = codes_declares - codes_ref
                    if hors_ref:
                        _gnoms_ui = _lire_global_noms(src_ref)
                        df_extra = pd.DataFrame([
                            {"CODE": c, "NOM": _gnoms_ui.get(c, "⚠️ Non répertoriée")}
                            for c in sorted(hors_ref)
                        ])
                        df_declares = pd.concat([df_declares, df_extra], ignore_index=True)
                else:
                    df_declares = pd.DataFrame([{"CODE": c, "NOM": "—"} for c in sorted(codes_declares)])

                df_declares = df_declares.sort_values("CODE").reset_index(drop=True)
                df_declares.index += 1
                st.dataframe(
                    df_declares,
                    use_container_width=True,
                    column_config={
                        "CODE": st.column_config.TextColumn("Code banque"),
                        "NOM" : st.column_config.TextColumn("Nom de la banque"),
                    },
                )
            else:
                st.info("Aucune banque valide détectée.")

        # ── 5b. Banques n'ayant PAS déclaré ───────────────────────────────────
        codes_non_declares = codes_ref - codes_declares
        with st.expander(f"❌ Banques n'ayant pas déclaré ({len(codes_non_declares)})", expanded=True):
            if not df_ref.empty and codes_non_declares:
                df_non_dec = df_ref[df_ref["CODE"].isin(codes_non_declares)].copy()
                df_non_dec = df_non_dec[["CODE", "NOM"]].sort_values("CODE").reset_index(drop=True)
                df_non_dec.index += 1
                st.dataframe(
                    df_non_dec,
                    use_container_width=True,
                    column_config={
                        "CODE": st.column_config.TextColumn("Code banque"),
                        "NOM" : st.column_config.TextColumn("Nom de la banque"),
                    },
                )
            elif df_ref.empty:
                st.warning("Fichier de référence introuvable ou vide — impossible de lister les non-déclarants.")
            else:
                st.success("✅ Toutes les banques répertoriées ont déclaré.")

        # ── 5c. Fichiers mal structurés ────────────────────────────────────────
        with st.expander(
            f"⚠️ Fichiers mal structurés / sans code banque valide ({len(resultat['fichiers_mal_struct'])})",
            expanded=len(resultat["fichiers_mal_struct"]) > 0,
        ):
            if resultat["fichiers_mal_struct"]:
                df_mal = pd.DataFrame(resultat["fichiers_mal_struct"])
                df_mal.index += 1
                st.dataframe(df_mal, use_container_width=True, column_config={
                    "Fichier": st.column_config.TextColumn("Fichier"),
                    "Raison" : st.column_config.TextColumn("Raison du rejet"),
                })
            else:
                st.success("✅ Aucun fichier mal structuré.")

        # ── 5d. Téléchargement du rapport CSV ─────────────────────────────────
        st.divider()
        if not df_ref.empty:
            # Construction du rapport complet
            lignes_rapport = []
            for _, row in df_ref.iterrows():
                statut = "✅ Déclarée" if row["CODE"] in codes_declares else "❌ Non déclarée"
                lignes_rapport.append({"CODE BANQUE": row["CODE"], "NOM BANQUE": row["NOM"], "STATUT": statut})
            for item in resultat["fichiers_mal_struct"]:
                lignes_rapport.append({
                    "CODE BANQUE": "—",
                    "NOM BANQUE" : item["Fichier"],
                    "STATUT"     : f"⚠️ Fichier rejeté : {item['Raison']}",
                })
            df_rapport = pd.DataFrame(lignes_rapport)

            csv_bytes = df_rapport.to_csv(index=False, sep=";", encoding="utf-8-sig").encode("utf-8-sig")
            st.download_button(
                label     = "⬇️ Télécharger le rapport CSV",
                data      = csv_bytes,
                file_name = f"Rapport_{type_dfx}_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                mime      = "text/csv",
                use_container_width=True,
            )


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN — NAVIGATION
# ══════════════════════════════════════════════════════════════════════════════



# ══════════════════════════════════════════════════════════════════════════════
#  MODULE 2 : TAUX DE RÉTROCESSION
# ══════════════════════════════════════════════════════════════════════════════

# Valeurs considérées comme XAF (monnaie locale) dans la colonne
# "Devise du compte du bénéficiaire" de la 1500M
DEVISES_XAF = {"XAF", "FRANC CFA", "FCFA", "CFA"}


def _normaliser_devise(v) -> str:
    return str(v).strip().upper() if v is not None else ""


def _normaliser_code(s) -> str:
    """
    Normalise un code banque en éliminant le suffixe flottant '.0' induit par Excel.
    Ex. : '10007.0' -> '10007'   |   'CITIGB2L' -> 'CITIGB2L'   |  10007.0 -> '10007'
    """
    s = str(s).strip()
    try:
        return str(int(float(s)))  # '10007.0' ou 10007 → '10007'
    except (ValueError, OverflowError):
        return s  # codes BIC ou autres non-numériques inchangés


# Valeurs parasites à ignorer dans les pivots
_VALS_PARASITES = {"TOTAL GÉNÉRAL", "TOTAL GENERAL", "(VIDE)", "#REF!", "NAN", ""}


def _normaliser_src(src):
    """
    Normalise une source de données : chemin str OU UploadedFile/BytesIO.
    Retourne un BytesIO (si fichier uploadé) ou un str (chemin disque).
    """
    if isinstance(src, str):
        return src.replace("\\", "/").strip()
    # UploadedFile Streamlit ou tout file-like
    return io.BytesIO(src.getvalue() if hasattr(src, "getvalue") else src.read())


def _construire_map_noms(src_ref) -> dict:
    """
    Lit le fichier de référence codes banques et retourne un dict {CODE: NOM}.
    Accepte un chemin str ou un UploadedFile.
    """
    if src_ref is None:
        return {}
    try:
        data = _normaliser_src(src_ref)
        if isinstance(data, str) and not os.path.exists(data):
            return {}
        # Essaie de lire la bonne feuille ; sinon prend la première
        xf = pd.ExcelFile(data)
        sheet = next(
            (s for s in xf.sheet_names if "banque" in s.lower() or "code" in s.lower()),
            xf.sheet_names[0]
        )
        df = pd.read_excel(xf, sheet_name=sheet, header=0, dtype=str).dropna(how="all")
        df = df.apply(lambda c: c.str.strip() if c.dtype == object else c)
        cols = list(df.columns)
        # Colonne code : première colonne contenant "code"
        c_code = next((c for c in cols if "code" in str(c).lower()), cols[0])
        # Colonne nom : première colonne avec libellé banque
        c_nom  = next(
            (c for c in cols if any(k in str(c).lower()
             for k in ("nom", "libel", "raison", "banque", "denom")) and c != c_code),
            cols[1] if len(cols) > 1 else None
        )
        if c_nom is None:
            return {}
        return dict(zip(
            df[c_code].fillna("").astype(str).str.strip().str.upper(),
            df[c_nom].fillna("").astype(str).str.strip()
        ))
    except Exception:
        return {}


def _extraire_noms_consolides(*sources) -> dict:
    """
    Extrait les noms de banques directement depuis les feuilles Déclarants ✅ /
    Non-déclarants ❌ des fichiers consolidés uploadés.
    Retourne un dict {CODE_UPPER: NOM}.
    """
    mapping = {}
    for src in sources:
        if src is None:
            continue
        try:
            data = _normaliser_src(src)
            xf = pd.ExcelFile(data)
            for sheet in xf.sheet_names:
                su = sheet.upper()
                if "CLARANT" not in su:   # filtre rapide : Déclarants / Non-déclarants
                    continue
                df = pd.read_excel(xf, sheet_name=sheet, header=0, dtype=str).dropna(how="all")
                cols = list(df.columns)
                c_code = next((c for c in cols if "CODE" in str(c).upper()), None)
                c_nom  = next(
                    (c for c in cols if any(k in str(c).upper()
                     for k in ("NOM", "LIBEL", "BANQUE")) and c != c_code),
                    None
                )
                if not c_code or not c_nom:
                    continue
                for _, row in df.iterrows():
                    code = str(row[c_code]).strip() if pd.notna(row[c_code]) else ""
                    nom  = str(row[c_nom]).strip()  if pd.notna(row[c_nom])  else ""
                    if code and nom and code.upper() not in ("NAN", "NONE", ""):
                        mapping[code.upper()] = nom
        except Exception:
            pass
    return mapping


def _lire_flat_consolide(src, label: str = "1200M",
                          col_montant_out: str = "MONTANT_1200M") -> pd.DataFrame | None:
    """
    Lit un fichier consolidé PLAT généré par l'appli (header=0).
    Détecte automatiquement les colonnes : CODE BANQUE | MONTANT EN DEVISE | DEVISE(S).
    Retourne : CODE_BANQUE | DEVISE | <col_montant_out>
    Accepte un chemin str ou un UploadedFile.
    """
    data = _normaliser_src(src)
    if isinstance(data, str) and not os.path.exists(data):
        return None
    try:
        df = pd.read_excel(data, header=0)
        cols = {str(c).strip().upper(): c for c in df.columns}
        c_code    = next((v for k, v in cols.items() if "CODE" in k and "BANQUE" in k), None)
        c_montant = next((v for k, v in cols.items() if "MONTANT" in k and "DEVISE" in k
                          and "FCFA" not in k), None)
        c_devise  = (next((v for k, v in cols.items() if k in ("DEVISE", "DEVISES")), None)
                     or next((v for k, v in cols.items()
                              if k.startswith("DEVISE") and "MONTANT" not in k
                              and "COMPTE" not in k and "OPERATION" not in k
                              and "OP" not in k), None))
        if not all([c_code, c_montant, c_devise]):
            st.warning(
                f"{label} — colonnes non trouvées.\n"
                f"  CODE BANQUE : {'\u2714' if c_code else '\u2718 absent'}  "
                f"  MONTANT : {'\u2714' if c_montant else '\u2718 absent'}  "
                f"  DEVISE : {'\u2714' if c_devise else '\u2718 absent'}\n"
                f"  Colonnes dispo : {list(df.columns)[:12]}"
            )
            return None
        df2 = df[[c_code, c_devise, c_montant]].copy()
        df2.columns = ["CODE_BANQUE", "DEVISE", "MONTANT"]
        df2["CODE_BANQUE"] = df2["CODE_BANQUE"].astype(str).str.strip().apply(_normaliser_code)
        df2["DEVISE"]      = df2["DEVISE"].astype(str).str.strip().str.upper()
        df2["MONTANT"]     = pd.to_numeric(df2["MONTANT"], errors="coerce").fillna(0)
        # Filtrer les lignes parasites
        df2 = df2[df2["MONTANT"] != 0]
        df2 = df2[~df2["CODE_BANQUE"].isin(["nan", "None", "", "NaN", "CODE BANQUE"])]
        df2 = df2[~df2["DEVISE"].isin(_VALS_PARASITES)]
        if df2.empty:
            st.warning(f"{label} — aucune ligne après filtrage (montant ≠ 0, code non vide).")
            return None
        return df2.groupby(["CODE_BANQUE", "DEVISE"], as_index=False)["MONTANT"].sum() \
                  .rename(columns={"MONTANT": col_montant_out})
    except Exception as e:
        st.error(f"Erreur lecture {label} : {e}")
        return None


def _lire_1200m_plat(src) -> pd.DataFrame | None:
    """
    Lit DFX_1200M_Consolide.xlsx ou DFX_1200Y_Consolide.xlsx (table plate générée par l'appli).
    Colonnes utiles : CODE BANQUE | MONTANT EN DEVISE | DEVISE(S)
    Retourne : CODE_BANQUE | DEVISE | MONTANT_1200M
    Accepte un chemin str ou un UploadedFile.
    """
    return _lire_flat_consolide(src, label="1200M", col_montant_out="MONTANT_1200M")


def _lire_1401m_pivot(src) -> pd.DataFrame | None:
    """
    Lit le fichier 1401M consolidé, en détectant automatiquement le format :
      - Format PLAT (produit par l'appli) : header=0, colonnes CODE BANQUE + MONTANT EN DEVISE + DEVISE
      - Format PIVOT (produit par Excel) : en-tête en ligne 4 (skiprows=3), colonnes = devises
    Retourne : CODE_BANQUE | DEVISE | MONTANT_1401M
    Accepte un chemin str ou un UploadedFile.
    """
    data = _normaliser_src(src)
    if isinstance(data, str) and not os.path.exists(data):
        return None
    try:
        # ── Détection du format : plat ou pivot ──────────────────────────────────
        if isinstance(data, io.BytesIO):
            data.seek(0)
        df_probe = pd.read_excel(data, header=0, nrows=3)
        if isinstance(data, io.BytesIO):
            data.seek(0)

        # Si la 1ère colonne contient "CODE" et "BANQUE" → format plat généré par l'appli
        first_col_upper = str(df_probe.columns[0]).strip().upper()
        is_flat = ("CODE" in first_col_upper and "BANQUE" in first_col_upper)

        if is_flat:
            # ── Lecture format PLAT ────────────────────────────────────────────
            return _lire_flat_consolide(data, label="1401M", col_montant_out="MONTANT_1401M")

        # ── Lecture format PIVOT (SWIFT_Consolide1401M.xlsx) ─────────────────────
        if isinstance(data, io.BytesIO):
            data.seek(0)
        df = pd.read_excel(data, skiprows=3, header=0)
        # Première colonne = code banque
        col_banque = df.columns[0]
        df.rename(columns={col_banque: "CODE_BANQUE"}, inplace=True)
        # Colonnes devises = tout sauf CODE_BANQUE, Total, (vide), Unnamed
        cols_dev = [
            c for c in df.columns
            if c != "CODE_BANQUE"
            and str(c).strip().upper() not in _VALS_PARASITES
            and not str(c).startswith("Unnamed")
            and "TOTAL" not in str(c).upper()
        ]
        df = df[["CODE_BANQUE"] + cols_dev].copy()
        df = df[df["CODE_BANQUE"].notna()]
        df = df[~df["CODE_BANQUE"].astype(str).str.contains(
            r"Total|Étiquettes|vide|#REF", case=False, na=False, regex=True)]
        df["CODE_BANQUE"] = df["CODE_BANQUE"].astype(str).str.strip().apply(_normaliser_code)
        df_long = df.melt(id_vars=["CODE_BANQUE"], value_vars=cols_dev,
                          var_name="DEVISE", value_name="MONTANT")
        df_long["DEVISE"]  = df_long["DEVISE"].astype(str).str.strip().str.upper()
        df_long["MONTANT"] = pd.to_numeric(df_long["MONTANT"], errors="coerce").fillna(0)
        df_long = df_long[~df_long["DEVISE"].isin(_VALS_PARASITES)]
        df_long = df_long[df_long["MONTANT"] != 0]
        return df_long.groupby(["CODE_BANQUE", "DEVISE"], as_index=False)["MONTANT"].sum() \
                      .rename(columns={"MONTANT": "MONTANT_1401M"})
    except Exception as e:
        st.error(f"Erreur lecture 1401M : {e}")
        return None


def _detecter_ligne_entete_1500m(src, max_scan: int = 20) -> int:
    """
    Scanne les premières lignes du fichier pour trouver la vraie ligne d'en-tête
    contenant 'Code banque' ou 'CODE BANQUE' avec le plus de colonnes renseignées.
    Accepte un chemin str ou un BytesIO.
    """
    # Pour les BytesIO, on lit puis on rewind
    if isinstance(src, io.BytesIO):
        src.seek(0)
        df_raw = pd.read_excel(src, header=None, nrows=max_scan)
        src.seek(0)
    else:
        df_raw = pd.read_excel(src, header=None, nrows=max_scan)
    candidats = []
    for i, row in df_raw.iterrows():
        n_vals = sum(1 for v in row.values if str(v).strip() not in ("nan", "None", ""))
        for val in row.values:
            s = str(val).strip().upper()
            if "CODE" in s and "BANQUE" in s:
                candidats.append((n_vals, int(i)))
                break
    if not candidats:
        return 0
    return max(candidats, key=lambda x: x[0])[1]


def _lire_1500m_plat(src, filtre: str = "tous") -> pd.DataFrame | None:
    """
    Lit le fichier 1500M plat (table complète), filtre sur la devise du compte,
    et retourne (CODE_BANQUE, DEVISE, MONTANT_1500M).
    filtre : 'xaf' | 'hors_xaf' | 'tous'
    Accepte un chemin str ou un UploadedFile.
    """
    data = _normaliser_src(src)
    if isinstance(data, str) and not os.path.exists(data):
        return None
    try:
        ligne_hdr = _detecter_ligne_entete_1500m(data)
        # Rewind BytesIO après détection
        if isinstance(data, io.BytesIO):
            data.seek(0)
        df = pd.read_excel(data, header=ligne_hdr)

        cols = {str(c).strip().upper(): c for c in df.columns}
        c_code        = next((v for k, v in cols.items() if "CODE" in k and "BANQUE" in k), None)
        c_devise_cpte = next((v for k, v in cols.items() if k.startswith("DEVISE") and "COMPTE" in k), None)
        c_montant     = next((v for k, v in cols.items() if k.startswith("MONTANT") and "DEVISE" in k
                              and "FCFA" not in k), None)
        # Devise de l'opération : peut s'appeler "Devise de l'opération" ou simplement "Devise"
        c_devise_op   = (next((v for k, v in cols.items() if k.startswith("DEVISE") and "OP" in k), None)
                         or next((v for k, v in cols.items()
                                  if k in ("DEVISE", "DEVISES") and v != c_devise_cpte), None)
                         or next((v for k, v in cols.items()
                                  if k.startswith("DEVISE") and "COMPTE" not in k
                                  and v != c_devise_cpte), None))
        if not all([c_code, c_montant, c_devise_op]):
            st.warning(f"1500M — colonnes non trouvées (en-tête détectée ligne {ligne_hdr + 1}). "
                       f"Dispo : {list(df.columns)[:10]}")
            return None
        df2 = df.copy()
        if filtre in ("xaf", "hors_xaf") and c_devise_cpte:
            mask = df2[c_devise_cpte].astype(str).str.strip().str.upper().isin(DEVISES_XAF)
            df2 = df2[mask] if filtre == "xaf" else df2[~mask]
        df2 = df2[[c_code, c_devise_op, c_montant]].copy()
        df2.columns = ["CODE_BANQUE", "DEVISE", "MONTANT"]
        df2["CODE_BANQUE"] = df2["CODE_BANQUE"].astype(str).str.strip().apply(_normaliser_code)
        df2["DEVISE"]      = df2["DEVISE"].astype(str).str.strip().str.upper()
        df2["MONTANT"]     = pd.to_numeric(df2["MONTANT"], errors="coerce").fillna(0)
        df2 = df2[df2["MONTANT"] != 0]
        return df2.groupby(["CODE_BANQUE", "DEVISE"], as_index=False)["MONTANT"].sum() \
                  .rename(columns={"MONTANT": "MONTANT_1500M"})
    except Exception as e:
        st.error(f"Erreur lecture 1500M : {e}")
        return None


def _calculer_taux(df_1500: pd.DataFrame,
                   df_autre: pd.DataFrame,
                   label_autre: str = "MONTANT_DFX") -> pd.DataFrame:
    """
    Joint df_1500 et df_autre sur (CODE_BANQUE, DEVISE),
    calcule Taux = (MONTANT_DFX / MONTANT_1500M) × 100.
    """
    df = pd.merge(df_1500, df_autre, on=["CODE_BANQUE", "DEVISE"], how="outer")
    df["MONTANT_1500M"]  = df["MONTANT_1500M"].fillna(0)
    df[label_autre]      = df[label_autre].fillna(0)

    df["TAUX (%)"] = df.apply(
        lambda r: round((r[label_autre] / r["MONTANT_1500M"]) * 100, 4)
        if r["MONTANT_1500M"] != 0 else 0.0,
        axis=1,
    )
    return df.sort_values(["CODE_BANQUE", "DEVISE"]).reset_index(drop=True)


def _pivoter(df: pd.DataFrame, col_val: str) -> pd.DataFrame:
    """Crée un tableau pivot CODE_BANQUE × DEVISE."""
    return df.pivot_table(index="CODE_BANQUE", columns="DEVISE",
                          values=col_val, aggfunc="sum", fill_value=0) \
             .reset_index()


def _df_to_excel_bytes(sheets: dict) -> bytes:
    """Crée un fichier Excel multi-feuilles en mémoire."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for nom, df in sheets.items():
            df.to_excel(writer, sheet_name=nom[:31], index=False)
            ws = writer.sheets[nom[:31]]
            # En-tête bleu
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF", size=10)
                cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78",
                                        fill_type="solid")
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
            ws.freeze_panes = "A2"
    buf.seek(0)
    return buf.getvalue()


def _ajouter_noms(df: pd.DataFrame, map_noms: dict) -> pd.DataFrame:
    """Insère la colonne NOM_BANQUE juste après CODE_BANQUE si map_noms est fourni."""
    if not map_noms:
        return df
    df = df.copy()
    df.insert(1, "NOM_BANQUE",
              df["CODE_BANQUE"].astype(str).str.upper().map(map_noms).fillna("—"))
    return df


def _afficher_tableau_avec_pivot(df_detail: pd.DataFrame,
                                  col_montant_1500: str,
                                  col_montant_dfx: str,
                                  titre: str,
                                  map_noms: dict = None):
    """Affiche le tableau détaillé + pivot des taux dans l'UI."""
    df_display = _ajouter_noms(df_detail, map_noms)

    tab1, tab2, tab3 = st.tabs(["Detail banque-devise",
                                  "Pivot Taux %",
                                  "Pivot Montants"])

    with tab1:
        try:
            fmt = {col_montant_1500: "{:,.2f}",
                   col_montant_dfx : "{:,.2f}",
                   "TAUX (%)"      : "{:.2f}"}
            st.dataframe(
                df_display.style.format(fmt, na_rep="-"),
                use_container_width=True,
                height=420,
            )
        except Exception:
            st.dataframe(df_display, use_container_width=True, height=420)

    with tab2:
        df_pivot_taux = _ajouter_noms(_pivoter(df_detail, "TAUX (%)"), map_noms)
        try:
            fmt2 = {c: "{:.2f}" for c in df_pivot_taux.columns
                    if c not in ("CODE_BANQUE", "NOM_BANQUE")}
            st.dataframe(
                df_pivot_taux.style.format(fmt2, na_rep="-"),
                use_container_width=True,
            )
        except Exception:
            st.dataframe(df_pivot_taux, use_container_width=True)

    with tab3:
        df_pivot_1500 = _ajouter_noms(_pivoter(df_detail, col_montant_1500), map_noms)
        df_pivot_dfx  = _ajouter_noms(_pivoter(df_detail, col_montant_dfx),  map_noms)
        st.write(f"**Montants 1500M filtre - {titre}**")
        try:
            st.dataframe(df_pivot_1500.style.format(
                {c: "{:,.2f}" for c in df_pivot_1500.columns
                 if c not in ("CODE_BANQUE", "NOM_BANQUE")},
                na_rep="-"
            ), use_container_width=True)
        except Exception:
            st.dataframe(df_pivot_1500, use_container_width=True)
        st.write(f"**Montants {col_montant_dfx}**")
        try:
            st.dataframe(df_pivot_dfx.style.format(
                {c: "{:,.2f}" for c in df_pivot_dfx.columns
                 if c not in ("CODE_BANQUE", "NOM_BANQUE")},
                na_rep="-"
            ), use_container_width=True)
        except Exception:
            st.dataframe(df_pivot_dfx, use_container_width=True)


def module_retrocession():
    try:
        _module_retrocession_inner()
    except Exception as exc:
        st.error(f"Erreur inattendue dans le module : {exc}")
        import traceback
        st.code(traceback.format_exc())


def _module_retrocession_inner():
    # ── Paramètres ─────────────────────────────────────────────────────────────
    with st.container(border=True):
        st.subheader("📥 Chargement des fichiers")
        st.caption(
            "Importez vos 3 fichiers consolidés ci-dessous. "
            "Le module accepte indifféremment les fichiers mensuels (M) et annuels (Y)."
        )

        # ── Bannière d'aide rapide ──────────────────────────────────────────
        with st.expander("ℹ️  Quels fichiers importer ?", expanded=False):
            st.markdown("""
| # | Fichier | Description | Exemples de noms |
|---|---------|-------------|-----------------|
| 1 | **DFX 1200M/Y — Consolidé** | Table plate des rétrocessions (devise XAF/FCFA) | `DFX_1200M_Consolide.xlsx`, `DFX_1200Y_Consolide.xlsx` |
| 2 | **DFX 1401M/Y — Consolidé** | Transferts en devises étrangères *(table plate **ou** pivot Excel)* | `DFX_1401M_Consolide.xlsx`, `SWIFT_Consolide1401M.xlsx` |
| 3 | **DFX 1500M/Y — Consolidé** | Tous les transferts — base de calcul des taux | `DFX_1500M_Consolide.xlsx`, `DFX_1500Y_Consolide.xlsx` |
            """)

        # ── 3 uploaders côte à côte ─────────────────────────────────────────
        col1, col2, col3 = st.columns(3)

        def _upload_card(col, numero, titre, sous_titre, badge_couleur, badge_texte, exemples):
            with col:
                statut_style = f"background:{badge_couleur};color:#fff;border-radius:4px;padding:1px 8px;font-size:11px;font-weight:700;"
                st.markdown(f"""
                <div style="background:#003087;border:2px solid #C8A951;border-radius:10px;
                            padding:14px 16px;margin-bottom:10px;min-height:110px;">
                    <div style="display:flex;align-items:center;gap:10px;margin-bottom:6px;">
                        <div style="background:#C8A951;color:#00205B;font-weight:900;font-size:18px;
                                    border-radius:50%;width:32px;height:32px;display:flex;
                                    align-items:center;justify-content:center;flex-shrink:0;">{numero}</div>
                        <div>
                            <div style="font-size:14px;font-weight:800;color:#FFFFFF;">{titre}
                                <span style="{statut_style}">{badge_texte}</span>
                            </div>
                            <div style="font-size:11.5px;color:#9FBBDE;margin-top:2px;">{sous_titre}</div>
                        </div>
                    </div>
                    <div style="font-size:11px;color:#C8D8F0;border-top:1px solid rgba(200,169,81,0.3);
                                padding-top:6px;margin-top:4px;">{exemples}</div>
                </div>
                """, unsafe_allow_html=True)

        _upload_card(col1, "1", "DFX 1200M / 1200Y", "Rétrocessions — devise FCFA/XAF",
                     "#1F6B3E", "OBLIGATOIRE",
                     "Ex. : <code>DFX_1200M_Consolide.xlsx</code> ou <code>DFX_1200Y_Consolide.xlsx</code>")
        _upload_card(col2, "2", "DFX 1401M / 1401Y", "Transferts devises étrangères",
                     "#1F6B3E", "OBLIGATOIRE",
                     "Ex. : <code>DFX_1401M_Consolide.xlsx</code> ou <code>SWIFT_Consolide1401M.xlsx</code>")
        _upload_card(col3, "3", "DFX 1500M / 1500Y", "Tous transferts — base de calcul",
                     "#1F6B3E", "OBLIGATOIRE",
                     "Ex. : <code>DFX_1500M_Consolide.xlsx</code> ou <code>DFX_1500Y_Consolide.xlsx</code>")

        with col1:
            up_1200m = st.file_uploader(
                "📂  Importer le fichier 1200M/Y",
                type=["xlsx", "xlsm"], key="up_1200m",
                help="Colonnes requises : CODE BANQUE | MONTANT EN DEVISE | DEVISE(S)",
            )
            if up_1200m: st.success(f"✅ {up_1200m.name}")
        with col2:
            up_1401m = st.file_uploader(
                "📂  Importer le fichier 1401M/Y",
                type=["xlsx", "xlsm"], key="up_1401m",
                help="Accepte table plate (app) ou pivot Excel (SWIFT_Consolide). "
                     "Colonnes requises : CODE BANQUE | MONTANT EN DEVISE | DEVISE",
            )
            if up_1401m: st.success(f"✅ {up_1401m.name}")
        with col3:
            up_1500m = st.file_uploader(
                "📂  Importer le fichier 1500M/Y",
                type=["xlsx", "xlsm"], key="up_1500m",
                help="Colonnes requises : CODE BANQUE | MONTANT EN DEVISE | DEVISE DE L'OPÉRATION | DEVISE DU COMPTE",
            )
            if up_1500m: st.success(f"✅ {up_1500m.name}")

        st.divider()
        col_ref, col_nom = st.columns([2, 1])
        with col_ref:
            st.markdown("""
            <div style="background:linear-gradient(135deg,#3D2B00,#2A1E00);border:2px solid #C8A951;
                        border-radius:8px;padding:10px 14px;margin-bottom:6px;">
            <b style="color:#FFD97E;">📌  Référentiel des noms de banques</b>
            <span style="color:#C8A951;font-size:11px;margin-left:8px;">(optionnel)</span><br>
            <span style="color:#D4B870;font-size:12px;">
            Enrichit les tableaux avec les noms complets des banques.<br>
            Ex. : <code style="color:#FFD97E;">Codes_banques_régul_Jan26.xlsx</code>
            </span>
            </div>
            """, unsafe_allow_html=True)
            up_ref = st.file_uploader(
                "📋  Fichier codes/noms banques (optionnel)",
                type=["xlsx", "xlsm"], key="up_ref_retro",
                help="Colonnes attendues : CODE BANQUE | NOM (ou libellé/dénomination).",
            )
            if up_ref: st.success(f"✅ Référentiel : {up_ref.name}")
        with col_nom:
            nom_sortie = st.text_input("Nom du fichier résultat", value="Taux_Retrocession.xlsx")

    # ── Résolution des sources ─────────────────────────────────────────────────
    src_1200 = up_1200m
    src_1401 = up_1401m
    src_1500 = up_1500m

    # Statut visuel récapitulatif
    st.divider()
    ok1 = up_1200m is not None
    ok2 = up_1401m is not None
    ok3 = up_1500m is not None

    def _badge_html(ok, label):
        bg = "#1F6B3E" if ok else "#8B0000"
        ic = "✅" if ok else "❌"
        return f'<span style="background:{bg};color:#fff;padding:4px 12px;border-radius:6px;font-weight:700;font-size:13px;">{ic} {label}</span>'

    st.markdown(
        f'<div style="display:flex;gap:16px;align-items:center;flex-wrap:wrap;padding:6px 0;">'
        f'{_badge_html(ok1, "1200M")} {_badge_html(ok2, "1401M")} {_badge_html(ok3, "1500M")}'
        f'<span style="color:#888;font-size:12px;margin-left:8px;">'
        f'{"→ Prêt à calculer !" if all([ok1,ok2,ok3]) else "→ Veuillez importer les 3 fichiers obligatoires"}'
        f'</span></div>',
        unsafe_allow_html=True
    )

    # ── Lancement ──────────────────────────────────────────────────────────────
    if st.button("🔢 Calculer les taux de rétrocession", type="primary",
                 use_container_width=True):

        if not all([ok1, ok2, ok3]):
            st.error("❌ Veuillez importer les 3 fichiers Excel avant de calculer.")
            return

        # Noms de banques : extraits automatiquement depuis les feuilles
        # Déclarants/Non-déclarants des consolidés, puis surchargés par le
        # référentiel optionnel (qui prend la priorité s'il est fourni).
        map_noms_consolide = _extraire_noms_consolides(up_1200m, up_1401m, up_1500m)
        map_noms_ref       = _construire_map_noms(up_ref) if up_ref is not None else {}
        map_noms           = {**map_noms_consolide, **map_noms_ref}

        # ── Lecture ────────────────────────────────────────────────────────────
        with st.spinner("Lecture du fichier 1200M/Y…"):
            df_1200 = _lire_1200m_plat(src_1200)

        with st.spinner("Lecture du fichier 1401M/Y (détection auto plat/pivot)…"):
            df_1401 = _lire_1401m_pivot(src_1401)

        with st.spinner("Lecture et filtrage du fichier 1500M/Y…"):
            df_1500_xaf  = _lire_1500m_plat(src_1500, filtre="xaf")
            df_1500_hors = _lire_1500m_plat(src_1500, filtre="hors_xaf")
            df_1500_tous = _lire_1500m_plat(src_1500, filtre="tous")

        for label, df in [("1200M/Y", df_1200), ("1401M/Y", df_1401),
                           ("1500M/Y(XAF)", df_1500_xaf), ("1500M/Y(HorsXAF)", df_1500_hors),
                           ("1500M/Y(Tous)", df_1500_tous)]:
            if df is None or df.empty:
                st.error(f"Impossible de lire / données vides : {label}")
                return

        # ── Calcul taux 1200M ─────────────────────────────────────────────────
        with st.spinner("Calcul taux 1200M…"):
            df_taux_1200 = _calculer_taux(df_1500_xaf, df_1200, "MONTANT_1200M")

        # ── Calcul taux 1401M ─────────────────────────────────────────────────
        with st.spinner("Calcul taux 1401M…"):
            df_taux_1401 = _calculer_taux(df_1500_hors, df_1401, "MONTANT_1401M")

        # ── Calcul taux Global ────────────────────────────────────────────────
        with st.spinner("Calcul taux global…"):
            df_global = pd.concat(
                [df_1200.rename(columns={"MONTANT_1200M": "MONTANT"}),
                 df_1401.rename(columns={"MONTANT_1401M": "MONTANT"})],
                ignore_index=True
            ).groupby(["CODE_BANQUE", "DEVISE"], as_index=False)["MONTANT"].sum()
            df_global.rename(columns={"MONTANT": "MONTANT_GLOBAL"}, inplace=True)
            df_taux_global = _calculer_taux(df_1500_tous, df_global, "MONTANT_GLOBAL")

        # ── Métriques ─────────────────────────────────────────────────────────
        st.divider()
        st.subheader("📊 Résultats")

        # Diagnostic codes correspondance ─────────────────────────────────────
        codes_1200 = set(df_1200["CODE_BANQUE"].unique())
        codes_1401 = set(df_1401["CODE_BANQUE"].unique())
        codes_1500 = set(df_1500_tous["CODE_BANQUE"].unique())
        codes_match_1200 = codes_1200 & codes_1500
        codes_match_1401 = codes_1401 & codes_1500

        if not codes_match_1200 and not codes_match_1401:
            st.warning(
                "⚠️ **Aucune correspondance de codes banques** entre les fichiers 1200M/1401M et 1500M. "
                "Les taux seront tous à 0. Vérifiez que vos fichiers utilisent le **même format de code** "
                "(numérique ex. `10007` ou BIC ex. `CITIGB2L`) dans les 3 fichiers."
            )
        elif len(codes_match_1200) < len(codes_1200) * 0.5 or len(codes_match_1401) < len(codes_1401) * 0.5:
            st.info(
                f"ℹ️ Correspondance partielle : {len(codes_match_1200)}/{len(codes_1200)} codes "
                f"1200M et {len(codes_match_1401)}/{len(codes_1401)} codes 1401M retrouvés dans 1500M."
            )

        nb_banques  = len(set(df_taux_1200["CODE_BANQUE"].tolist() +
                              df_taux_1401["CODE_BANQUE"].tolist()))
        tmoy_1200   = df_taux_1200[df_taux_1200["TAUX (%)"] > 0]["TAUX (%)"].mean()
        tmoy_1401   = df_taux_1401[df_taux_1401["TAUX (%)"] > 0]["TAUX (%)"].mean()
        tmoy_global = df_taux_global[df_taux_global["TAUX (%)"] > 0]["TAUX (%)"].mean()

        k1, k2, k3, k4 = st.columns(4)
        k1.metric("Banques (codes uniques)", nb_banques)
        k2.metric("Taux moyen 1200M/Y", f"{tmoy_1200:.2f} %"  if not pd.isna(tmoy_1200)   else "—")
        k3.metric("Taux moyen 1401M/Y", f"{tmoy_1401:.2f} %"  if not pd.isna(tmoy_1401)   else "—")
        k4.metric("Taux moyen global",   f"{tmoy_global:.2f} %" if not pd.isna(tmoy_global) else "—")

        if map_noms:
            src_label = "référentiel + consolidés" if map_noms_ref else "feuilles Déclarants/Non-déclarants des consolidés"
            st.caption(f"✅ {len(map_noms)} noms de banques chargés depuis les {src_label}.")

        st.divider()

        with st.expander("📋 Taux DFX 1200M/Y  —  base 1500M filtre XAF / FRANC CFA", expanded=True):
            st.caption("Filtre 1500M : Devise du compte ∈ {XAF, FRANC CFA, FCFA}. "
                       "Taux = (Montant 1200M / Montant 1500M filtré) × 100")
            _afficher_tableau_avec_pivot(df_taux_1200, "MONTANT_1500M", "MONTANT_1200M",
                                         "1200M", map_noms)

        with st.expander("📋 Taux DFX 1401M/Y  —  base 1500M filtre HORS XAF", expanded=True):
            st.caption("Filtre 1500M : Devise du compte ∉ {XAF, FRANC CFA, FCFA}. "
                       "Taux = (Montant 1401M / Montant 1500M filtré) × 100")
            _afficher_tableau_avec_pivot(df_taux_1401, "MONTANT_1500M", "MONTANT_1401M",
                                         "1401M", map_noms)

        with st.expander("📊 Taux Global  —  (1200M/Y + 1401M/Y) / 1500M sans filtre", expanded=True):
            st.caption("Aucun filtre sur la 1500M. "
                       "Taux = (Montant 1200M + Montant 1401M) / Montant 1500M total × 100")
            _afficher_tableau_avec_pivot(df_taux_global, "MONTANT_1500M", "MONTANT_GLOBAL",
                                         "Global", map_noms)

        # ── Export ────────────────────────────────────────────────────────────
        st.divider()
        # Enrichir avec noms pour l'export aussi
        def _enrichi(df): return _ajouter_noms(df, map_noms)
        excel_bytes = _df_to_excel_bytes({
            "Taux 1200M"          : _enrichi(df_taux_1200),
            "Taux 1401M"          : _enrichi(df_taux_1401),
            "Taux Global"         : _enrichi(df_taux_global),
            "Pivot Taux 1200M"    : _pivoter(df_taux_1200,   "TAUX (%)"),
            "Pivot Taux 1401M"    : _pivoter(df_taux_1401,   "TAUX (%)"),
            "Pivot Taux Global"   : _pivoter(df_taux_global,  "TAUX (%)"),
            "Agg 1200M"           : df_1200,
            "Agg 1401M"           : df_1401,
            "Agg 1500M (XAF)"     : df_1500_xaf,
            "Agg 1500M (HorsXAF)" : df_1500_hors,
            "Agg 1500M (Tous)"    : df_1500_tous,
        })
        st.download_button(
            label     = f"⬇️ Télécharger {nom_sortie}",
            data      = excel_bytes,
            file_name = nom_sortie,
            mime      = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
        st.success(
            f"✅ Calculs terminés — {len(df_taux_1200)} lignes 1200M/Y | "
            f"{len(df_taux_1401)} lignes 1401M/Y | {len(df_taux_global)} lignes Global"
        )


# ══════════════════════════════════════════════════════════════════════════════
#  MODULE 3 : DOMICILIATIONS DES EXPORTATIONS
# ══════════════════════════════════════════════════════════════════════════════

def _dom_export_code_banque(ws):
    """
    Extrait le code banque d'un fichier DOM EXPORT.
    - Feuille INFOS : lit directement la cellule E5.
    - Autres feuilles : essaie D3, C3, E3, D5,
      puis recherche l'étiquette 'Code Ban*' dans les 6 premières lignes.
    """
    # Feuille INFOS → code en E5
    if ws.title.strip().upper() == "INFOS":
        val = ws["E5"].value
        if val is not None and str(val).strip():
            try:
                return str(int(float(str(val).strip())))
            except (ValueError, OverflowError):
                s = str(val).strip()
                if re.match(r"^\d{4,8}$", s):
                    return s

    # Fallback sur des cellules fixes
    for cell_ref in ("D3", "C3", "E3", "D5"):
        val = ws[cell_ref].value
        if val is not None and str(val).strip():
            try:
                return str(int(float(str(val).strip())))
            except (ValueError, OverflowError):
                s = str(val).strip()
                if re.match(r"^\d{4,8}$", s):
                    return s

    # Recherche étiquetée dans les 6 premières lignes
    for row_idx in range(1, 7):
        row_vals = [ws.cell(row=row_idx, column=col).value for col in range(1, 12)]
        for i, cell_val in enumerate(row_vals):
            if cell_val and "code ban" in str(cell_val).lower():
                for j in range(i + 1, min(i + 5, len(row_vals))):
                    adj = row_vals[j]
                    if adj is not None and str(adj).strip():
                        try:
                            return str(int(float(str(adj).strip())))
                        except (ValueError, OverflowError):
                            return str(adj).strip()
    return None


def _dom_export_annee(sheet_name: str):
    """
    Extrait l'année (4 chiffres) depuis le nom d'une feuille DOM EXPORT.
    Exemples : '2022' → '2022', 'NOV.25' → '2025', 'Domiciliations 2023' → '2023'.
    Retourne None si aucune année identifiable.
    """
    s = sheet_name.strip()
    # Cas 1 : nom = exactement 4 chiffres
    if re.fullmatch(r"\d{4}", s):
        return s
    # Cas 2 : contient une année 4 chiffres (ex. 'Domiciliations 2022')
    m = re.search(r"\b(20\d{2})\b", s)
    if m:
        return m.group(1)
    # Cas 3 : format MOIS.AA ex. 'NOV.25', 'DEC.25'
    m = re.match(r"[A-Za-z\u00C0-\u00FF]+\.(\d{2})$", s)
    if m:
        return str(2000 + int(m.group(1)))
    return None


def _dom_export_find_header(ws):
    """
    Détecte automatiquement la ligne d'en-tête en cherchant le mot 'exportateur'
    (couvre 'Nom de l\'exportateur' et 'Nom Exportateur').
    Retourne (header_row_1based, col_nom_0based, col_facture_0based, col_rapatrmt_0based).
    """
    for row_idx in range(1, 20):
        max_col = min(ws.max_column, 30)
        row_vals = [ws.cell(row=row_idx, column=col).value for col in range(1, max_col + 1)]
        row_strs = [str(v).lower().strip() if v is not None else "" for v in row_vals]
        if any("exportateur" in s for s in row_strs):
            col_nom, col_facture, col_rapatrmt = None, None, None
            for i, s in enumerate(row_strs):
                if col_nom is None and "exportateur" in s:
                    col_nom = i
                if col_facture is None and "facture" in s and "montant" in s:
                    col_facture = i
                if col_rapatrmt is None and ("règlement" in s or "regleme" in s or "rapatriem" in s) \
                        and "montant" in s:
                    col_rapatrmt = i
            return row_idx, col_nom, col_facture, col_rapatrmt
    return None, None, None, None


def concatener_dom_export(fichiers) -> dict:
    """
    Concatène les fichiers DOM EXPORT (domiciliations exportations).
    Pour chaque fichier, parcourt les feuilles dont le nom = année (4 chiffres).
    Extrait 4 colonnes : Code Banque, Nom Exportateur, Montant Facture, Montant Rapatriement.
    Retourne { 'dfs_annee': {str: DataFrame}, 'fichiers_ok', 'fichiers_erreur', 'total_lignes' }.
    """
    COLONNES = ["Code Banque", "Nom Exportateur", "Montant Facture", "Montant Rapatriement"]
    donnees_par_annee: dict = {}
    fichiers_ok = []
    fichiers_erreur = []
    total_lignes = 0

    for f in fichiers:
        nom = f.name
        try:
            wb = openpyxl.load_workbook(io.BytesIO(f.getvalue()), data_only=True)
            # Code banque depuis la feuille INFOS (cellule E5)
            infos_name = next((s for s in wb.sheetnames if s.strip().upper() == "INFOS"), None)
            code_banque_global = (
                _dom_export_code_banque(wb[infos_name]) if infos_name else None
            )

            lignes_fichier = 0
            traite_ok = False

            for sheet_name in wb.sheetnames:
                # Ignorer la feuille INFOS
                if sheet_name.strip().upper() == "INFOS":
                    continue

                annee = _dom_export_annee(sheet_name)
                if annee is None:
                    continue  # Feuille sans année identifiable

                ws = wb[sheet_name]

                # Code banque depuis cette feuille si non trouvé dans INFOS
                code_banque = code_banque_global or _dom_export_code_banque(ws) or "—"

                header_row, col_nom, col_facture, col_rapatrmt = _dom_export_find_header(ws)
                if header_row is None or col_nom is None:
                    continue

                max_row = ws.max_row
                max_col = min(ws.max_column, 30)

                # Lire les lignes de données brutes
                data_rows_raw = [
                    [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
                    for r in range(header_row + 1, max_row + 1)
                ]

                # Correction auto : si col_nom contient des entiers (N° séquence), décaler
                premiers_vals = [
                    row[col_nom] for row in data_rows_raw[:5]
                    if col_nom < len(row) and row[col_nom] is not None
                ]
                if (premiers_vals
                        and all(isinstance(v, (int, float)) for v in premiers_vals)
                        and col_nom + 1 < max_col):
                    shifted = [
                        row[col_nom + 1] for row in data_rows_raw[:5]
                        if col_nom + 1 < len(row) and row[col_nom + 1] is not None
                    ]
                    if shifted and any(isinstance(v, str) for v in shifted):
                        col_nom += 1

                for row_vals in data_rows_raw:
                    if col_nom >= len(row_vals):
                        continue
                    nom_exp   = row_vals[col_nom] if col_nom is not None else None
                    mt_fact   = row_vals[col_facture]  if (col_facture  is not None and col_facture  < len(row_vals)) else None
                    mt_rapatr = row_vals[col_rapatrmt] if (col_rapatrmt is not None and col_rapatrmt < len(row_vals)) else None

                    if nom_exp is None or str(nom_exp).strip() == "":
                        continue
                    if mt_fact is None and mt_rapatr is None:
                        continue

                    donnees_par_annee.setdefault(annee, []).append({
                        "Code Banque"        : code_banque,
                        "Nom Exportateur"    : str(nom_exp).strip(),
                        "Montant Facture"    : mt_fact,
                        "Montant Rapatriement": mt_rapatr,
                    })
                    lignes_fichier += 1
                    traite_ok = True

            total_lignes += lignes_fichier
            if traite_ok:
                fichiers_ok.append(nom)
            else:
                fichiers_erreur.append({"Fichier": nom, "Raison": "Aucune donnée extraite (vérifier structure)"})
            wb.close()

        except Exception as exc:
            fichiers_erreur.append({"Fichier": nom, "Raison": str(exc)})

    dfs_annee = {}
    for annee, lignes in sorted(donnees_par_annee.items()):
        df = pd.DataFrame(lignes, columns=COLONNES)
        df["Montant Facture"]      = pd.to_numeric(df["Montant Facture"],      errors="coerce")
        df["Montant Rapatriement"] = pd.to_numeric(df["Montant Rapatriement"], errors="coerce")
        dfs_annee[annee] = df
    return {
        "dfs_annee"      : dfs_annee,
        "fichiers_ok"    : fichiers_ok,
        "fichiers_erreur": fichiers_erreur,
        "total_lignes"   : total_lignes,
    }


def _construire_dom_export_excel(dfs_annee: dict, df_ref: pd.DataFrame = None) -> bytes:
    """
    Construit le fichier Excel DOM EXPORT en mémoire.
    Feuilles : une par année + 'Top 10 Exportateurs' + 'Top 10 Banques'
               + 'Rapport Déclarations' + 'Taux par Banque'.
    df_ref : DataFrame avec colonnes CODE, NOM (référentiel des banques).
    """
    COLONNES = ["Code Banque", "Nom Exportateur", "Montant Facture", "Montant Rapatriement"]
    LARGEURS  = [16, 42, 24, 26]

    hdr_font  = Font(bold=True, size=11, color="FFFFFF")
    hdr_fill  = PatternFill(start_color="003087", end_color="003087", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border    = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    def _write_sheet(ws_out, df, cols, widths):
        for ci, col_name in enumerate(cols, 1):
            c = ws_out.cell(row=1, column=ci, value=col_name)
            c.font = hdr_font; c.fill = hdr_fill
            c.alignment = hdr_align; c.border = border
        for ri, row_data in enumerate(df.itertuples(index=False), 2):
            for ci, val in enumerate(row_data, 1):
                c = ws_out.cell(row=ri, column=ci, value=val)
                c.border = border
                if isinstance(val, (int, float)) and ci >= 3:
                    c.number_format = "#,##0.00"
                    c.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    c.alignment = Alignment(horizontal="left", vertical="center")
        ws_out.freeze_panes = "A2"
        if ws_out.dimensions and ws_out.dimensions != "A1:A1":
            ws_out.auto_filter.ref = ws_out.dimensions
        for ci, w in enumerate(widths, 1):
            ws_out.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    all_dfs = list(dfs_annee.values())
    df_all  = pd.concat(all_dfs, ignore_index=True) if all_dfs else pd.DataFrame(columns=COLONNES)

    # Feuilles par année
    for annee, df in dfs_annee.items():
        ws_an = wb.create_sheet(f"Données {annee}"[:31])
        _write_sheet(ws_an, df, COLONNES, LARGEURS)

    # Top 10 Exportateurs
    ws_exp = wb.create_sheet("Top 10 Exportateurs")
    if not df_all.empty:
        df_top_exp = (
            df_all.groupby("Nom Exportateur", dropna=True)["Montant Facture"]
            .sum()
            .reset_index()
            .rename(columns={"Montant Facture": "Volume total (Facture)"})
            .sort_values("Volume total (Facture)", ascending=False)
            .head(10)
            .reset_index(drop=True)
        )
        df_top_exp.insert(0, "Rang", range(1, len(df_top_exp) + 1))
        _write_sheet(ws_exp, df_top_exp, list(df_top_exp.columns), [8, 44, 26])
    else:
        ws_exp.cell(row=1, column=1, value="Aucune donnée disponible")

    # Top 10 Banques
    ws_bq = wb.create_sheet("Top 10 Banques")
    if not df_all.empty:
        df_top_bq = (
            df_all.groupby("Code Banque", dropna=True)["Montant Facture"]
            .sum()
            .reset_index()
            .rename(columns={"Montant Facture": "Volume total (Facture)"})
            .sort_values("Volume total (Facture)", ascending=False)
            .head(10)
            .reset_index(drop=True)
        )
        df_top_bq.insert(0, "Rang", range(1, len(df_top_bq) + 1))
        _write_sheet(ws_bq, df_top_bq, list(df_top_bq.columns), [8, 20, 26])
    else:
        ws_bq.cell(row=1, column=1, value="Aucune donnée disponible")

    # ── Rapport Déclarations ──────────────────────────────────────────────────
    ws_decl = wb.create_sheet("Rapport Déclarations")
    codes_declares = set(df_all["Code Banque"].dropna().astype(str).str.strip().unique()) if not df_all.empty else set()

    lignes_rapport = []
    if df_ref is not None and not df_ref.empty:
        codes_ref = set(df_ref["CODE"].astype(str).str.strip())
        # Banques du référentiel : déclarées ou non
        for _, row in df_ref.iterrows():
            code = str(row["CODE"]).strip()
            nom  = str(row["NOM"]).strip()
            statut = "✅ Déclarée" if code in codes_declares else "❌ Non déclarée"
            lignes_rapport.append({"Code Banque": code, "Nom Banque": nom, "Statut": statut})
        # Banques hors référentiel mais ayant quand même déposé un fichier
        hors_ref = codes_declares - codes_ref
        for code in sorted(hors_ref):
            lignes_rapport.append({"Code Banque": code, "Nom Banque": "⚠️ Hors référentiel", "Statut": "✅ Déclarée"})
    else:
        # Pas de référentiel : on liste uniquement les banques ayant déposé
        for code in sorted(codes_declares):
            lignes_rapport.append({"Code Banque": code, "Nom Banque": "—", "Statut": "✅ Déclarée"})

    if lignes_rapport:
        df_rapport = pd.DataFrame(lignes_rapport).sort_values(["Statut", "Code Banque"], ascending=[False, True]).reset_index(drop=True)

        # Couleurs selon statut
        fill_dec     = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # vert pâle
        fill_non_dec = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # rouge pâle
        font_dec     = Font(bold=False, color="276221")
        font_non_dec = Font(bold=False, color="9C0006")

        # En-tête
        hdr_cols_decl = ["Code Banque", "Nom Banque", "Statut"]
        for ci, col_name in enumerate(hdr_cols_decl, 1):
            c = ws_decl.cell(row=1, column=ci, value=col_name)
            c.font = hdr_font; c.fill = hdr_fill
            c.alignment = hdr_align; c.border = border

        for ri, row_data in enumerate(df_rapport.itertuples(index=False), 2):
            is_declared = str(row_data[2]).startswith("✅")
            row_fill = fill_dec if is_declared else fill_non_dec
            row_font = font_dec if is_declared else font_non_dec
            for ci, val in enumerate(row_data, 1):
                c = ws_decl.cell(row=ri, column=ci, value=val)
                c.border = border
                c.fill = row_fill
                c.font = row_font
                c.alignment = Alignment(horizontal="left", vertical="center")

        ws_decl.freeze_panes = "A2"
        ws_decl.auto_filter.ref = f"A1:C{len(df_rapport) + 1}"
        for ci, w in enumerate([16, 42, 18], 1):
            ws_decl.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
    else:
        ws_decl.cell(row=1, column=1, value="Aucune donnée disponible")

    # ── Taux par Banque ───────────────────────────────────────────────────────
    ws_taux = wb.create_sheet("Taux par Banque")
    if not df_all.empty and "Montant Facture" in df_all.columns and "Montant Rapatriement" in df_all.columns:
        df_taux = (
            df_all.groupby("Code Banque", dropna=True)
            .agg(
                Total_Facture     =("Montant Facture",       "sum"),
                Total_Rapatriement=("Montant Rapatriement",  "sum"),
            )
            .reset_index()
        )
        # Taux = Montant Facture / Montant Rapatriement (0 si dénominateur nul)
        df_taux["Taux (Facture/Rapatrié)"] = df_taux.apply(
            lambda r: (r["Total_Facture"] / r["Total_Rapatriement"])
            if r["Total_Rapatriement"] and r["Total_Rapatriement"] != 0
            else None,
            axis=1,
        )
        df_taux = df_taux.rename(columns={
            "Total_Facture"     : "Montant Total Facture",
            "Total_Rapatriement": "Montant Total Rapatrié",
        }).sort_values("Montant Total Facture", ascending=False).reset_index(drop=True)

        # Si référentiel disponible, enrichir avec le nom de la banque
        if df_ref is not None and not df_ref.empty:
            map_noms = dict(zip(df_ref["CODE"].astype(str).str.strip(), df_ref["NOM"].astype(str).str.strip()))
            df_taux.insert(1, "Nom Banque", df_taux["Code Banque"].map(map_noms).fillna("—"))
            cols_taux  = ["Code Banque", "Nom Banque", "Montant Total Facture", "Montant Total Rapatrié", "Taux (Facture/Rapatrié)"]
            widths_taux = [16, 38, 26, 26, 24]
        else:
            cols_taux  = ["Code Banque", "Montant Total Facture", "Montant Total Rapatrié", "Taux (Facture/Rapatrié)"]
            widths_taux = [16, 26, 26, 24]

        # En-tête
        for ci, col_name in enumerate(cols_taux, 1):
            c = ws_taux.cell(row=1, column=ci, value=col_name)
            c.font = hdr_font; c.fill = hdr_fill
            c.alignment = hdr_align; c.border = border

        for ri, row_data in enumerate(df_taux[cols_taux].itertuples(index=False), 2):
            for ci, val in enumerate(row_data, 1):
                c = ws_taux.cell(row=ri, column=ci, value=val)
                c.border = border
                col_name = cols_taux[ci - 1]
                if col_name == "Taux (Facture/Rapatrié)":
                    c.number_format = "0.0000"
                    c.alignment = Alignment(horizontal="right", vertical="center")
                elif "Montant" in col_name:
                    c.number_format = "#,##0.00"
                    c.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    c.alignment = Alignment(horizontal="left", vertical="center")

        ws_taux.freeze_panes = "A2"
        ws_taux.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(cols_taux))}{len(df_taux) + 1}"
        for ci, w in enumerate(widths_taux, 1):
            ws_taux.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
    else:
        ws_taux.cell(row=1, column=1, value="Aucune donnée disponible")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def module_dom_export():
    """Module 3 : Domiciliations des exportations."""

    # ── Paramètres ─────────────────────────────────────────────────────────────
    st.subheader("⚙️ Paramètres")
    col_a, col_b = st.columns([3, 2])
    with col_b:
        nom_sortie = st.text_input(
            "Nom du fichier de sortie",
            value=f"DOM_EXPORT_Consolide_{datetime.now().strftime('%Y%m')}.xlsx",
            key="dom_nom_sortie",
        )

    st.info(
        "📂 **Fichiers DOM EXPORT** — Importez un ou plusieurs fichiers Excel "
        "(une banque par fichier). Les feuilles nommées par année (**2024, 2025…**) "
        "sont traitées automatiquement. Colonnes retenues : "
        "**Code Banque** (D3) · **Nom Exportateur** · **Montant Facture** · **Montant Rapatriement**"
    )

    fichiers_dom = st.file_uploader(
        "📂  Glissez-déposez vos fichiers DOM EXPORT ici, ou cliquez pour parcourir",
        type=["xlsx", "xlsm"],
        accept_multiple_files=True,
        key="up_fichiers_dom",
        help="Ex. : CBC_CMR_02_2026.xlsx, NFC_CM_02_2026_B_EXPORT.xlsx",
    )
    if fichiers_dom:
        noms  = ", ".join(f.name for f in fichiers_dom[:4])
        extra = f" (+{len(fichiers_dom) - 4} autres)" if len(fichiers_dom) > 4 else ""
        st.success(f"✅  **{len(fichiers_dom)} fichier(s) chargé(s)** : {noms}{extra}")
    else:
        st.warning("⚠️  Aucun fichier sélectionné — importez vos fichiers DOM EXPORT pour continuer.")

    st.divider()

    if st.button("🚀 Lancer la concaténation DOM EXPORT", type="primary", use_container_width=True):
        if not fichiers_dom:
            st.error("❌ Veuillez importer au moins un fichier DOM EXPORT.")
            return

        with st.spinner(f"Traitement de {len(fichiers_dom)} fichier(s) en cours…"):
            resultat = concatener_dom_export(fichiers_dom)

        # Persister dans session_state pour survivre aux re-renders suivants
        st.session_state["dom_resultat"] = resultat

    # ── Affichage des résultats (depuis session_state) ──────────────────────
    if "dom_resultat" not in st.session_state:
        return

    resultat        = st.session_state["dom_resultat"]
    dfs_annee       = resultat["dfs_annee"]
    total_lignes    = resultat["total_lignes"]
    fichiers_ok     = resultat["fichiers_ok"]
    fichiers_erreur = resultat["fichiers_erreur"]

    if total_lignes == 0:
        st.error(
            "❌ Aucune donnée extraite. Vérifiez que vos fichiers ont des feuilles "
            "nommées par année et la colonne « Nom de l'exportateur »."
        )
        if fichiers_erreur:
            st.dataframe(pd.DataFrame(fichiers_erreur), use_container_width=True)
        return

    # ── Métriques ──────────────────────────────────────────────────────────────
    m1, m2, m3 = st.columns(3)
    m1.metric("Fichiers traités", len(fichiers_ok))
    m2.metric("Lignes extraites", f"{total_lignes:,}")
    m3.metric("Années détectées", len(dfs_annee))

    # ── Aperçu par année ───────────────────────────────────────────────────────
    if dfs_annee:
        st.subheader("📋 Aperçu des données par année")
        for annee, df in sorted(dfs_annee.items()):
            with st.expander(f"Année {annee} — {len(df):,} lignes", expanded=True):
                st.dataframe(df.head(30), use_container_width=True)

    # ── Tops ───────────────────────────────────────────────────────────────────
    df_all = pd.concat(list(dfs_annee.values()), ignore_index=True) if dfs_annee else pd.DataFrame()
    if not df_all.empty and "Montant Facture" in df_all.columns:
        st.subheader("🏆 Classements par volume (Montant Facture)")
        col_top1, col_top2 = st.columns(2)

        with col_top1:
            st.markdown("**Top 10 Exportateurs**")
            df_top_exp = (
                df_all.groupby("Nom Exportateur", dropna=True)["Montant Facture"]
                .sum()
                .reset_index()
                .rename(columns={"Montant Facture": "Volume Facture"})
                .sort_values("Volume Facture", ascending=False)
                .head(10)
                .reset_index(drop=True)
            )
            df_top_exp.insert(0, "Rang", range(1, len(df_top_exp) + 1))
            st.dataframe(df_top_exp, use_container_width=True)

        with col_top2:
            st.markdown("**Top 10 Banques**")
            df_top_bq = (
                df_all.groupby("Code Banque", dropna=True)["Montant Facture"]
                .sum()
                .reset_index()
                .rename(columns={"Montant Facture": "Volume Facture"})
                .sort_values("Volume Facture", ascending=False)
                .head(10)
                .reset_index(drop=True)
            )
            df_top_bq.insert(0, "Rang", range(1, len(df_top_bq) + 1))
            st.dataframe(df_top_bq, use_container_width=True)

    # ── Fichiers en erreur ─────────────────────────────────────────────────────
    if fichiers_erreur:
        with st.expander(f"⚠️ Fichiers avec erreurs ({len(fichiers_erreur)})", expanded=True):
            st.dataframe(pd.DataFrame(fichiers_erreur), use_container_width=True)

    # ── Téléchargement ─────────────────────────────────────────────────────────
    st.divider()

    # Construire df_ref depuis le référentiel global (toutes feuilles de Codes_banques.xlsx)
    df_ref_dom = pd.DataFrame(columns=["CODE", "NOM"])
    try:
        if os.path.exists(FICHIER_CODES_BANQUES):
            map_noms_global = _lire_global_noms(FICHIER_CODES_BANQUES)
            if map_noms_global:
                df_ref_dom = pd.DataFrame(
                    [{"CODE": k, "NOM": v} for k, v in map_noms_global.items()]
                )
    except Exception:
        pass  # référentiel facultatif — ne bloque pas le téléchargement

    # ── Rapport déclarations (aperçu UI) ──────────────────────────────────────
    if dfs_annee:
        df_all_decl = pd.concat(list(dfs_annee.values()), ignore_index=True) if dfs_annee else pd.DataFrame()
        codes_declares_dom = set(df_all_decl["Code Banque"].dropna().astype(str).str.strip().unique()) if not df_all_decl.empty else set()

        st.subheader("📋 Rapport de déclaration — DOM EXPORT")
        lignes_ui = []
        if not df_ref_dom.empty:
            codes_ref_dom = set(df_ref_dom["CODE"].astype(str).str.strip())
            map_noms_ui   = dict(zip(df_ref_dom["CODE"].astype(str).str.strip(), df_ref_dom["NOM"].astype(str).str.strip()))
            for code in sorted(codes_ref_dom):
                statut = "✅ Déclarée" if code in codes_declares_dom else "❌ Non déclarée"
                lignes_ui.append({"Code Banque": code, "Nom Banque": map_noms_ui.get(code, "—"), "Statut": statut})
            for code in sorted(codes_declares_dom - codes_ref_dom):
                lignes_ui.append({"Code Banque": code, "Nom Banque": "⚠️ Hors référentiel", "Statut": "✅ Déclarée"})
        else:
            for code in sorted(codes_declares_dom):
                lignes_ui.append({"Code Banque": code, "Nom Banque": "—", "Statut": "✅ Déclarée"})

        if lignes_ui:
            df_rapport_ui = pd.DataFrame(lignes_ui).sort_values(["Statut", "Code Banque"], ascending=[False, True]).reset_index(drop=True)
            nb_dec     = (df_rapport_ui["Statut"] == "✅ Déclarée").sum()
            nb_non_dec = (df_rapport_ui["Statut"] == "❌ Non déclarée").sum()
            col_r1, col_r2 = st.columns(2)
            col_r1.metric("Banques ayant déclaré", nb_dec)
            col_r2.metric("Banques n'ayant pas déclaré", nb_non_dec)
            with st.expander("Voir le détail", expanded=True):
                st.dataframe(df_rapport_ui, use_container_width=True)

    xlsx_bytes = _construire_dom_export_excel(dfs_annee, df_ref=df_ref_dom if not df_ref_dom.empty else None)
    st.download_button(
        label            = f"⬇️ Télécharger {nom_sortie}",
        data             = xlsx_bytes,
        file_name        = nom_sortie,
        mime             = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN — NAVIGATION
# ══════════════════════════════════════════════════════════════════════════════

def main():
    st.set_page_config(
        page_title = "App Supervision DFX — BEAC",
        page_icon  = "🏦",
        layout     = "wide",
        initial_sidebar_state="expanded",
    )

    # Injection de la charte graphique BEAC
    _inject_beac_css()

    # ── Garde d'authentification ───────────────────────────────────────────────
    if not st.session_state.get("authenticated", False):
        _page_login()
        return

    # ── Changement de mot de passe obligatoire ────────────────────────────────
    if st.session_state.get("must_change_pwd", False):
        with st.sidebar:
            st.markdown("""
            <div style="text-align:center;padding:12px 0 4px;">
                <div style="font-size:18px;font-weight:800;color:#E8C96B;
                            letter-spacing:2px;">BEAC</div>
            </div>""", unsafe_allow_html=True)
        _page_changer_password()
        return

    # ── Récupération du profil courant ────────────────────────────────────────
    role          = st.session_state.get("role", "analyste_dfx")
    display_name  = st.session_state.get("display_name", "Utilisateur")
    modules_dispo = ROLES_MODULES.get(role, [])
    role_label    = ROLES_LABELS.get(role, role)
    ico, bg, fg   = ROLES_BADGES.get(role, ("👤", "#888", "#FFF"))

    # ── Sidebar ────────────────────────────────────────────────────────────────
    with st.sidebar:
        # Logo BEAC (emoji de substitution si le fichier n'existe pas)
        _logo = os.path.join(WORKSPACE, "logo_beac.png")
        if os.path.exists(_logo):
            st.image(_logo, use_container_width=True)
        else:
            st.markdown(
                '<div style="text-align:center;font-size:52px;padding:10px 0;">🏦</div>',
                unsafe_allow_html=True,
            )
        st.markdown("""
        <div style="text-align:center; padding: 4px 0 8px 0;">
            <div style="font-size:18px; font-weight:800; color:#E8C96B;
                        letter-spacing:2px; text-transform:uppercase;">BEAC</div>
            <div style="font-size:10px; color:#AAC0E0; letter-spacing:1px;
                        text-transform:uppercase; margin-top:2px;">
                Banque des États de l'Afrique Centrale
            </div>
        </div>
        """, unsafe_allow_html=True)

        # Carte profil utilisateur
        st.markdown(f"""
        <div style="background:rgba(0,48,135,0.25);border-radius:10px;
                    padding:10px 14px;margin:6px 0 10px;
                    border:1px solid rgba(200,169,81,0.3);">
            <div style="font-size:13px;font-weight:700;color:#FFFFFF;
                        margin-bottom:4px;">👤 {display_name}</div>
            <div>
                <span style="background:{bg};color:{fg};
                             padding:2px 10px;border-radius:12px;
                             font-size:11px;font-weight:700;">
                    {ico} {role_label}
                </span>
            </div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("""
        <div style="border-top:1px solid rgba(200,169,81,0.4);
                    border-bottom:1px solid rgba(200,169,81,0.4);
                    padding:6px 0; margin:8px 0; text-align:center;">
            <span style="font-size:12px; color:#C8D8F0; letter-spacing:0.8px;">
                Supervision des déclarations DFX
            </span>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        module = st.radio(
            "Navigation",
            options=modules_dispo,
            label_visibility="collapsed",
        )

        # Gestion des utilisateurs (admin uniquement)
        if role == "admin":
            st.markdown("<br>", unsafe_allow_html=True)
            _sidebar_gestion_users()

        st.markdown("<br>" * 2, unsafe_allow_html=True)
        st.divider()
        st.markdown(f"""
        <div style="font-size:10px; color:#8AA0C0; text-align:center; line-height:1.8;">
            <div style="margin-top:4px; color:#6080A0;">
                {datetime.now().strftime('%d/%m/%Y')}
            </div>
        </div>
        """, unsafe_allow_html=True)

        # Bouton de déconnexion
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🚪 Se déconnecter", use_container_width=True):
            for key in ["authenticated", "username", "display_name",
                        "role", "must_change_pwd"]:
                st.session_state.pop(key, None)
            st.session_state["_do_logout"] = True

    # Déclenchement du rerun HORS contexte sidebar (évite l'erreur removeChild)
    if st.session_state.pop("_do_logout", False) or st.session_state.pop("_gestion_rerun", False):
        st.rerun()

    # ── Bandeau de titre principal ─────────────────────────────────────────────
    if module == "Concaténation DFX":
        st.markdown("""
        <div style="
            background: linear-gradient(135deg, #003087 0%, #00205B 60%, #00153D 100%);
            border-radius: 12px;
            padding: 22px 30px;
            margin-bottom: 24px;
            border-left: 6px solid #C8A951;
            box-shadow: 0 4px 18px rgba(0,48,135,0.18);">
            <div style="display:flex; align-items:center; gap:16px;">
                <span style="font-size:40px;">📂</span>
                <div>
                    <div style="font-size:22px; font-weight:800; color:#FFFFFF;
                                letter-spacing:0.5px;">
                        Concaténation des déclarations DFX
                    </div>
                    <div style="font-size:13px; color:#C8D8F0; margin-top:4px;">
                        Consolidation automatique des fichiers DFX 1200M, 1401M et 1500M
                    </div>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)
        module_concatenation()

    elif module == "Taux de rétrocession":
        st.markdown("""
        <div style="
            background: linear-gradient(135deg, #003087 0%, #00205B 60%, #00153D 100%);
            border-radius: 12px;
            padding: 22px 30px;
            margin-bottom: 24px;
            border-left: 6px solid #C8A951;
            box-shadow: 0 4px 18px rgba(0,48,135,0.18);">
            <div style="display:flex; align-items:center; gap:16px;">
                <span style="font-size:40px;">📊</span>
                <div>
                    <div style="font-size:22px; font-weight:800; color:#FFFFFF;
                                letter-spacing:0.5px;">
                        Taux de rétrocession par banque
                    </div>
                    <div style="font-size:13px; color:#C8D8F0; margin-top:4px;">
                        Calcul des taux 1200M, 1401M et Global à partir des fichiers consolidés
                    </div>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)
        module_retrocession()

    elif module == "Domiciliations Export":
        st.markdown("""
        <div style="
            background: linear-gradient(135deg, #003087 0%, #00205B 60%, #00153D 100%);
            border-radius: 12px;
            padding: 22px 30px;
            margin-bottom: 24px;
            border-left: 6px solid #C8A951;
            box-shadow: 0 4px 18px rgba(0,48,135,0.18);">
            <div style="display:flex; align-items:center; gap:16px;">
                <span style="font-size:40px;">🚢</span>
                <div>
                    <div style="font-size:22px; font-weight:800; color:#FFFFFF;
                                letter-spacing:0.5px;">
                        Domiciliations des exportations
                    </div>
                    <div style="font-size:13px; color:#C8D8F0; margin-top:4px;">
                        Concaténation par année — Top 10 exportateurs et banques par volume
                    </div>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)
        module_dom_export()


if __name__ == "__main__":
    main()
